"""
Excel模板生成脚本 - 完整版本

生成项目数据Excel模板文件，包含全部6章数据结构。
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 获取项目根目录
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_DIR = os.path.join(PROJECT_ROOT, "templates", "excel_templates")


def apply_header_style(cell):
    """应用表头样式"""
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


def apply_cell_style(cell):
    """应用单元格样式"""
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def create_project_info_sheet(wb: Workbook):
    """创建项目基本信息Sheet（第1章）"""
    ws = wb.active
    ws.title = "项目基本信息"

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60

    # 表头
    ws['A1'] = "字段名称"
    ws['B1'] = "内容"
    apply_header_style(ws['A1'])
    apply_header_style(ws['B1'])

    # 数据
    data = [
        ("项目名称", "汉川市万福低闸等3座灌溉闸站更新改造工程项目"),
        ("项目代码", "2512-420984-04-01-395957"),
        ("建设单位", "汉川市水利和湖泊局"),
        ("建设性质", "更新改造"),
        ("项目投资", "7847.03万元"),
        ("项目选址", "龚家湾低闸泵站位于脉旺镇,万福低闸泵站、杜公泵站位于沉湖镇"),
        ("建设内容", "新建万福低闸泵站和龚家湾低闸泵站,改造杜公泵站。其中万福低闸灌溉泵站装机功率1350kW,设计流量10.0m³/s;龚家湾低闸灌溉泵站装机功率1800kW,设计流量12.0m³/s;杜公泵站装机功率560kW,设计流量3.0m³/s。"),
        ("建设规模", "总装机功率3710kW,总设计流量25.0m³/s"),
        ("建设期限", "24个月"),
        ("选址原则", "符合规划要求,不占优质耕地,尽量不迁移民,避免敏感区域,基础设施优先,集约节约利用,方便施工运营,安全可靠"),
    ]

    for i, (field, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=field)
        ws.cell(row=i, column=2, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 30


def create_alternatives_sheet(wb: Workbook):
    """创建备选方案Sheet（第2章）"""
    ws = wb.create_sheet("备选方案")

    # 设置列宽
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']
    widths = [10, 25, 25, 12, 15, 15, 15, 15, 12, 12, 12, 30, 20, 15, 15, 15, 15]
    for col, width in zip(columns, widths):
        ws.column_dimensions[col].width = width

    # 表头
    headers = [
        "方案编号", "方案名称", "位置", "面积(平方米)",
        "东", "南", "西", "北",
        "农村道路", "林地", "园地", "建设用地",
        "是否占用耕地", "是否占用永久基本农田", "是否涉及未利用地",
        "建设内容", "备注"
    ]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 示例数据 - 方案一
    row1 = [
        "1", "方案一：脉旺镇龚家湾", "汉川市脉旺镇龚家湾村", 10633.00,
        "农田", "沟渠", "村庄道路", "农田",
        "548.00", "8094.00", "1934.00", "57.00",
        "否", "否", "否",
        "新建龚家湾低闸灌溉泵站，装机功率1800kW，设计流量12.0m³/s", ""
    ]
    for i, value in enumerate(row1, start=1):
        cell = ws.cell(row=2, column=i, value=value)
        apply_cell_style(cell)

    # 示例数据 - 方案二
    row2 = [
        "2", "方案二：沉湖镇万福闸", "汉川市沉湖镇万福闸村", 10276.98,
        "农田", "河流", "村庄道路", "村庄",
        "0", "0", "8884.29", "1392.68",
        "否", "否", "否",
        "新建万福低闸灌溉泵站，装机功率1350kW，设计流量10.0m³/s", ""
    ]
    for i, value in enumerate(row2, start=1):
        cell = ws.cell(row=3, column=i, value=value)
        apply_cell_style(cell)

    for row in range(1, 4):
        ws.row_dimensions[row].height = 30


def create_site_conditions_sheet(wb: Workbook):
    """创建场地条件Sheet（第2章）"""
    ws = wb.create_sheet("场地条件")

    # 设置列宽
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50

    # 表头
    headers = ["类别", "项目", "内容"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 数据
    data = [
        # 地形地貌
        ("地形地貌", "地貌类型", "平原"),
        ("地形地貌", "地势走向", "地势平坦"),
        ("地形地貌", "最高海拔", "30米"),
        ("地形地貌", "最低海拔", "25米"),
        ("地形地貌", "对项目影响", "影响较小"),
        # 气候
        ("气候", "气候类型", "亚热带季风气候"),
        ("气候", "年均气温", "16-18℃"),
        ("气候", "年降雨量", "1200毫米"),
        ("气候", "极端最高温", "40℃"),
        ("气候", "极端最低温", "-10℃"),
        ("气候", "风速", "3.0m/s"),
        ("气候", "对施工影响", "影响较小"),
        # 区域地质构造
        ("区域地质构造", "构造位置", "江汉平原"),
        ("区域地质构造", "主要断裂", "无"),
        ("区域地质构造", "地壳稳定性", "稳定"),
        ("区域地质构造", "抗震设防烈度", "6度"),
        # 水文地质条件
        ("水文地质条件", "主要水系", "汉江"),
        ("水文地质条件", "河流长度", "无主要河流"),
        ("水文地质条件", "流域面积", "无"),
        ("水文地质条件", "年均产水量", "无"),
        # 工程地质
        ("工程地质", "岩组类型", "第四系松散土工程地质岩组"),
        ("工程地质", "承载力特征值", "150kPa"),
        ("工程地质", "岩土性质", "稳定性良好"),
        # 地震
        ("地震", "地震动峰值加速度", "0.05g"),
        ("地震", "地震动反应周期", "0.35s"),
        ("地震", "地震基本烈度", "6度"),
        ("地震", "引用标准", "《中国地震动参数区划图》（GB18306-2015）"),
        # 外部配套条件
        ("外部配套条件", "周边建筑物", "无影响施工的建筑物"),
        ("外部配套条件", "供水", "完备"),
        ("外部配套条件", "供电", "完备"),
        ("外部配套条件", "通讯", "完备"),
        ("外部配套条件", "交通", "省道，交通便利"),
        ("外部配套条件", "建材来源", "当地建材厂家供应充足"),
        ("外部配套条件", "是否压覆文物", "否"),
        ("外部配套条件", "是否影响防洪", "否"),
        # 规划影响
        ("规划影响", "是否符合国土空间总体规划", "是"),
        ("规划影响", "是否列入重点项目库", "是"),
        ("规划影响", "重点项目库名称", "水利发展重点项目库"),
        ("规划影响", "对区域发展作用", "促进农业发展，保障粮食安全"),
    ]

    for i, (category, item, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=item)
        ws.cell(row=i, column=3, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))
        apply_cell_style(ws.cell(row=i, column=3))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 25


def create_sensitive_sheet(wb: Workbook):
    """创建敏感条件Sheet（第2章）"""
    ws = wb.create_sheet("敏感条件")

    # 设置列宽
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40

    # 表头
    headers = ["类别", "项目", "内容"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 数据
    data = [
        # 历史保护
        ("历史保护", "是否压占历史文化名城", "否"),
        ("历史保护", "是否有古建筑古村落", "否"),
        ("历史保护", "是否有文物古迹", "否"),
        # 生态保护
        ("生态保护", "是否涉及自然保护区", "否"),
        ("生态保护", "是否有珍稀动植物", "否"),
        ("生态保护", "是否影响生态环境", "影响较小，已采取措施"),
        # 矿产资源
        ("矿产资源", "是否压覆矿产资源", "否"),
        ("矿产资源", "是否与采矿权重叠", "否"),
        ("矿产资源", "是否与探矿权重叠", "否"),
        # 安全防护
        ("安全防护", "是否满足邻避要求", "是"),
        ("安全防护", "是否涉及饮用水源保护区", "否"),
        ("安全防护", "是否涉及机场净空", "否"),
        ("安全防护", "是否涉及军事设施", "否"),
        # 重要设施
        ("重要设施", "机场", "无"),
        ("重要设施", "铁路", "距离较远，无影响"),
        ("重要设施", "公路", "省道，无影响"),
        ("重要设施", "水运", "不涉及"),
        # 耕地和基本农田
        ("耕地和基本农田", "是否占用耕地", "否"),
        ("耕地和基本农田", "是否占用永久基本农田", "否"),
        # 生态保护红线
        ("生态保护红线", "是否占用生态保护红线", "否"),
    ]

    for i, (category, item, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=item)
        ws.cell(row=i, column=3, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))
        apply_cell_style(ws.cell(row=i, column=3))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 25


def create_construction_sheet(wb: Workbook):
    """创建施工运营Sheet（第2章）"""
    ws = wb.create_sheet("施工运营")

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50

    # 表头
    ws['A1'] = "项目"
    ws['B1'] = "内容"
    apply_header_style(ws['A1'])
    apply_header_style(ws['B1'])

    # 数据
    data = [
        ("方案一总投资", "4500万元"),
        ("方案二总投资", "5200万元"),
        ("政府支持", "各级政府支持"),
        ("群众支持", "群众支持，征地拆迁已达成一致"),
        ("征地拆迁", "已完成征地拆迁协议签订"),
        ("施工难度", "较小"),
        ("材料供应", "充足"),
    ]

    for i, (field, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=field)
        ws.cell(row=i, column=2, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 30


def create_consultation_sheet(wb: Workbook):
    """创建征求意见Sheet（第2章）"""
    ws = wb.create_sheet("征求意见")

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 30

    # 表头
    headers = ["部门", "日期", "复函标题", "结论"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 数据
    data = [
        ("自然资源和规划局", "2025年9月3日", "《关于查询项目是否位于地质灾害易发区的复函》", "不属于地质灾害易发区"),
        ("文物保护站", "2025年9月3日", "《关于是否压占文物保护区的函》", "未发现地面文物"),
        ("生态环境局", "2025年9月4日", "《关于是否占用饮用水源保护地的复函》", "不涉及饮用水源保护地"),
        ("水利局", "2025年9月5日", "《关于项目建设对防洪影响的复函》", "对防洪无影响"),
        ("交通运输局", "2025年9月5日", "《关于项目建设对交通影响的复函》", "对交通无影响"),
    ]

    for i, row_data in enumerate(data, start=2):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            apply_cell_style(cell)

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 30


def create_comparison_sheet(wb: Workbook):
    """创建方案比选Sheet（第2章）"""
    ws = wb.create_sheet("方案比选")

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60

    # 表头
    ws['A1'] = "项目"
    ws['B1'] = "内容"
    apply_header_style(ws['A1'])
    apply_header_style(ws['B1'])

    # 数据
    data = [
        ("比选因子", "场地自然条件,外部配套条件,选址敏感条件,施工运营条件,规划影响条件"),
        ("推荐方案", "方案一"),
        ("推荐理由", "投资较低,交通条件更好,不占耕地和基本农田,有效避让生态保护红线"),
        ("比选原则", "科学性原则,经济性原则,可行性原则,可持续性原则"),
    ]

    for i, (field, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=field)
        ws.cell(row=i, column=2, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 30


# ============================================================================
# 第3章：合法合规性分析 Sheet
# ============================================================================

def create_regulation_sheet(wb: Workbook):
    """创建法规政策Sheet（第3章）"""
    ws = wb.create_sheet("法规政策")

    # 设置列宽
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15

    # 表头
    headers = ["法规名称", "发布单位", "发布时间", "符合性分析", "符合性结论"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 数据
    data = [
        ("《产业结构调整指导目录（2024年本）》", "国家发展和改革委员会", "2024年", 
         "项目属于第一类鼓励类：四十二、环境保护与资源节约综合利用中的3.城镇污水垃圾处理，符合国家产业政策。", "符合"),
        ("《划拨用地目录》", "国土资源部", "2001年",
         "项目符合城市基础设施用地和公益事业用地（三）城市基础设施用地5.环境卫生设施：包括污水处理厂，符合供地政策。", "符合"),
        ("《城镇污水处理厂污染物排放标准》（GB18918-2002）", "国家环境保护总局", "2002年",
         "出水水质满足一级A标准。", "符合"),
        ("《城市污水处理工程项目建设标准》（JB198-2022）", "住房和城乡建设部", "2022年",
         "项目用地规模符合建设标准要求。", "符合"),
    ]

    for i, row_data in enumerate(data, start=2):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            apply_cell_style(cell)

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 40


def create_three_lines_sheet(wb: Workbook):
    """创建三线分析Sheet（第3章）"""
    ws = wb.create_sheet("三线分析")

    # 设置列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    # 表头
    ws['A1'] = "项目"
    ws['B1'] = "内容"
    apply_header_style(ws['A1'])
    apply_header_style(ws['B1'])

    # 数据
    data = [
        ("是否占用耕地", "否"),
        ("占用耕地面积（平方米）", "0"),
        ("是否占用永久基本农田", "否"),
        ("占用永久基本农田面积（平方米）", "0"),
        ("是否占用生态保护红线", "否"),
        ("占用生态保护红线面积（平方米）", "0"),
        ("是否位于城镇开发边界内", "否"),
        ("城镇开发边界说明", "项目位于城镇开发边界外"),
        ("符合性说明", "项目属于《省自然资源厅关于加强'三区三线'实施管理的意见》规定的单独选址项目用地清单中民生基础设施项目，符合单独选址要求。拟选址项目区不涉及压占耕地、不涉及压占永久基本农田，不涉及压占生态保护红线，不位于城镇开发边界内。"),
        ("数据来源", "2023年国土变更调查数据"),
    ]

    for i, (field, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=field)
        ws.cell(row=i, column=2, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 30


def create_spatial_planning_sheet(wb: Workbook):
    """创建国土空间规划Sheet（第3章）"""
    ws = wb.create_sheet("国土空间规划")

    # 设置列宽
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50

    # 表头
    headers = ["类别", "项目", "内容"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 数据
    data = [
        # 一张图落位
        ("一张图落位", "是否上图落位", "是"),
        ("一张图落位", "重点项目库名称", "生态修复重点工程"),
        ("一张图落位", "项目类型", "生态修复类项目"),
        ("一张图落位", "落位说明", "项目选址方案符合国土空间总体规划，本项目在国土空间规划'一张图'上，已列入规划生态修复重点工程。"),
        # 功能分区准入
        ("功能分区准入", "城镇建设适宜性", "城镇建设适宜区"),
        ("功能分区准入", "生态保护重要性", "不属于生态保护极重要区"),
        ("功能分区准入", "农业生产适宜性", "农业生产不适宜区"),
        ("功能分区准入", "符合性说明", "项目选址方案在城镇适宜性评价结果图中属于城镇建设适宜区，在生态保护重要性评价结果图中不属于生态保护极重要区，于农业生产适宜性评价结果图中属于农业生产不适宜区，完全符合生态重要性等级、农业生产适宜等级和建设开发适宜等级等方面的符合性。"),
        # 用途管制
        ("用途管制", "符合性说明", "项目符合国土空间用途管制相关要求，用地性质为排水用地，符合规划用途。"),
        # 总体格局
        ("总体格局", "符合性说明", "项目整体符合国土空间总体规划的布局要求，已列入规划生态修复重点工程。"),
        # 总体结论
        ("总体结论", "总体符合性结论", "项目与国土空间总体规划符合性较高。"),
    ]

    for i, (category, item, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=item)
        ws.cell(row=i, column=3, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))
        apply_cell_style(ws.cell(row=i, column=3))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 30


def create_special_planning_sheet(wb: Workbook):
    """创建专项规划Sheet（第3章）"""
    ws = wb.create_sheet("专项规划")

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 15

    # 表头
    headers = ["规划类型", "规划名称", "符合性分析", "符合性结论"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 数据
    data = [
        ("综合交通规划", "兴山县国土空间总体规划（2021-2035年）-综合交通规划",
         "项目拟选址位于规划'四纵三横两支'的公路骨架网络的'横三'312省道沿线，该项目未占用S312省道沿线，并沿S312省道沿线一侧预留绿化带。", "符合"),
        ("市政基础设施规划", "兴山县国土空间总体规划（2021-2035年）-市政基础设施规划",
         "项目拟选址与市政基础设施规划中能源供应体系、通讯网络体系、水循环利用与污水治理体系相关规划均不冲突。", "符合"),
        ("历史文化遗产保护规划", "兴山县国土空间总体规划（2021-2035年）-历史文化遗产保护规划",
         "项目所在区域无压占历史文化名城、名镇、名村保护范围等，均未发现古城镇、古建筑、古村落，文物古迹保护等保护对象。", "符合"),
        ("综合防灾工程规划", "兴山县国土空间总体规划（2021-2035年）-综合防灾减灾体系",
         "项目所在区域位于一般防治区，主要面临轻度地质灾害和气象灾害风险，不与防灾重点工程冲突。", "符合"),
        ("旅游规划", "兴山县国土空间总体规划（2021-2035年）-城镇空间布局结构图",
         "项目所在区域位于峡口旅游物流组团片区内部，周边为香溪河绿色山水风景廊道轴线，不与现有旅游规划项目冲突，项目建成有助于维护周边环境。", "符合"),
        ("环境保护规划", "宜昌市环境保护总体规划（2013-2030年）",
         "项目拟选址位于生态黄线区，主要为小型点状开发的污水处理项目，针对生活污水处理加强了生态治理和修复。", "符合"),
        ("自然保护地规划", "湖北三峡万朝山省级自然保护区",
         "项目用地不在湖北三峡万朝山省级自然保护区范围内。", "符合"),
    ]

    for i, row_data in enumerate(data, start=2):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            apply_cell_style(cell)

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 40


def create_other_planning_sheet(wb: Workbook):
    """创建其他规划Sheet（第3章）"""
    ws = wb.create_sheet("其他规划")

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 15

    # 表头
    headers = ["规划类型", "规划名称", "符合性分析", "符合性结论"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 数据
    data = [
        ("国民经济和社会发展规划", "《兴山县国民经济和社会发展第十四个五年规划和2035年远景目标纲要》",
         "项目可对香溪河左岸峡口片区的生活污水进行系统治理和循环利用，与《规划纲要》中'持续抓好香溪河流域生态保护和修复'要点符合。", "符合"),
        ("生态环境保护规划", "《宜昌市环境保护总体规划（2013-2030年）》",
         "项目拟选址位于生态黄线区，符合管控要求。", "符合"),
        ("三线一单生态环境分区管控", "《宜昌市'三线一单'生态环境分区管控实施方案》",
         "项目拟选址位于重点管控单元，编号为ZH42052620003，项目不属于该单元空间布局约束禁止项目，符合管控单元污染物排放和环境风险防控相关要求。", "符合"),
        ("综合交通体系规划", "《宜昌市综合交通体系规划（2011-2030年）》",
         "项目拟选址位于规划的二级公路S312省道沿线，该项目未占用S312省道沿线，并沿S312省道沿线一侧预留绿化带。", "符合"),
    ]

    for i, row_data in enumerate(data, start=2):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            apply_cell_style(cell)

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 40


def create_urban_planning_sheet(wb: Workbook):
    """创建城乡总体规划Sheet（第3章）"""
    ws = wb.create_sheet("城乡总体规划")

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50

    # 表头
    ws['A1'] = "项目"
    ws['B1'] = "内容"
    apply_header_style(ws['A1'])
    apply_header_style(ws['B1'])

    # 数据
    data = [
        ("规划名称", "《宜昌市兴山县峡口镇总体规划（2014-2030）》"),
        ("规划期限", "2014-2030年"),
        ("空间管制分区", "适建区"),
        ("符合性分析", "项目拟选址位于镇域空间管制分类中适建区，适建区主要为工程地质条件良好、地势相对平坦、没有其它建设限制条件的区域。本项目建设规模、开发强度均较小，符合所在地经批复的城市总体规划布局的要求。"),
        ("符合性结论", "符合"),
    ]

    for i, (field, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=field)
        ws.cell(row=i, column=2, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 30


def create_compliance_summary_sheet(wb: Workbook):
    """创建合法合规小结Sheet（第3章）"""
    ws = wb.create_sheet("合法合规小结")

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80

    # 表头
    ws['A1'] = "项目"
    ws['B1'] = "内容"
    apply_header_style(ws['A1'])
    apply_header_style(ws['B1'])

    # 数据
    ws.cell(row=2, column=1, value="合法合规小结")
    ws.cell(row=2, column=2, value="综合前述项目与相关法律法规、政策文件的符合性分析、与'三线'和耕地等各类空间的协调分析、与国土空间总体规划的符合性分析、与专项规划的符合性分析以及与其他相关规划的符合性分析，建设项目总体上属于合法合规。")
    apply_cell_style(ws.cell(row=2, column=1))
    apply_cell_style(ws.cell(row=2, column=2))

    for row in range(1, 3):
        ws.row_dimensions[row].height = 60


# ============================================================================
# 第4章：选址合理性分析 Sheet
# ============================================================================

def create_environmental_sheet(wb: Workbook):
    """创建环境影响Sheet（第4章）"""
    ws = wb.create_sheet("环境影响")

    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 60

    # 表头
    headers = ["类别", "项目", "内容"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 数据
    data = [
        # 大气环境
        ("大气环境", "施工期扬尘措施1", "施工场地定期洒水，防止产生大量扬尘"),
        ("大气环境", "施工期扬尘措施2", "避免在春季大风季节以及夏季暴雨时节施工"),
        ("大气环境", "施工期扬尘措施3", "加强施工区的规划管理，建筑材料堆场采取防尘措施"),
        ("大气环境", "施工机械废气措施", "运输车辆严禁超载运输，加强对施工机械、车辆的维修保养"),
        ("大气环境", "运营期废气措施", "厨房安装油烟净化处理装置，引至中控楼顶高空排放"),
        ("大气环境", "影响程度", "影响较小"),
        ("大气环境", "防治结论", "采取上述措施后，大气环境影响可控，满足《环境空气质量标准》要求"),
        # 噪声环境
        ("噪声环境", "施工期噪声措施1", "将低噪声、低振动施工设备作为中标的重要内容"),
        ("噪声环境", "施工期噪声措施2", "设专人对施工设备进行定期保养和维护"),
        ("噪声环境", "施工期噪声措施3", "施工尽量安排在白天进行，尽量缩短工期"),
        ("噪声环境", "影响程度", "影响较小"),
        ("噪声环境", "防治结论", "施工区域距离声环境敏感目标较远，采取上述措施后可满足《建筑施工场界噪声限值》要求"),
        # 水环境
        ("水环境", "施工期废水措施", "施工生产废水集中收集处理，在场区内设置化粪池进行处理"),
        ("水环境", "运营期废水措施", "生活污水经化粪池处理后，基本接近《农田灌溉水质标准》"),
        ("水环境", "影响程度", "影响较小"),
        ("水环境", "防治结论", "项目建设对水环境影响较小，污水处理后达标排放"),
        # 固体废弃物
        ("固体废弃物", "施工期固废措施", "开挖土方时洒水降尘，注意土方临时堆放，剩余弃渣可作为场区附近低洼地段的填土"),
        ("固体废弃物", "影响程度", "影响较小"),
        ("固体废弃物", "防治结论", "固体废弃物得到妥善处置，对环境影响较小"),
        # 交通影响
        ("交通影响", "施工期交通影响", "工程建设时由于车辆运输可能导致交通暂时繁忙"),
        ("交通影响", "施工期缓解措施", "运输车辆宜避开交通高峰期，合理安排施工机械流程"),
        ("交通影响", "防治结论", "交通影响是暂时的，随着工程结束而消失"),
        # 生态修复
        ("生态修复", "对居民点影响", "项目地处区域距离居民点较远，对周边居民产生的影响甚微"),
        ("生态修复", "对动物影响", "项目施工期的基础建设、运输和设备安装对野生动物种群影响范围很小，施工结束后野生动物种群可逐渐迁回"),
        ("生态修复", "对植物影响", "永久占地原有生物量较小，场址范围内没有珍稀植物，建设对当地植物总体影响不大"),
        ("生态修复", "水土保持措施", "道路施工防治区设置边坡防护、截排水等工程防护措施"),
        # 环境影响小结
        ("环境影响小结", "小结内容", "项目区地质稳定，发生灾害可能性较小，自然条件良好，未压覆现有已探明矿产，对项目区周边环境影响较小"),
    ]

    for i, (category, item, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=item)
        ws.cell(row=i, column=3, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))
        apply_cell_style(ws.cell(row=i, column=3))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 25


def create_mineral_sheet(wb: Workbook):
    """创建矿产资源Sheet（第4章）"""
    ws = wb.create_sheet("矿产资源")

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60

    # 表头
    ws['A1'] = "项目"
    ws['B1'] = "内容"
    apply_header_style(ws['A1'])
    apply_header_style(ws['B1'])

    # 数据
    data = [
        ("是否压覆矿产资源", "否"),
        ("是否与采矿权重叠", "否"),
        ("是否与探矿权重叠", "否"),
        ("是否与地质项目重叠", "否"),
        ("复函信息", "兴山县自然资源和城乡建设局发布《关于项目用地是否压覆已查明矿产资源的复函》，确定项目用地范围及外扩300m不与采矿权、探矿权存在交叉重叠"),
        ("分析结论", "项目不存在压覆矿产资源的情况"),
    ]

    for i, (field, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=field)
        ws.cell(row=i, column=2, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 30


def create_geological_sheet(wb: Workbook):
    """创建地质灾害Sheet（第4章）"""
    ws = wb.create_sheet("地质灾害")

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60

    # 表头
    ws['A1'] = "项目"
    ws['B1'] = "内容"
    apply_header_style(ws['A1'])
    apply_header_style(ws['B1'])

    # 数据
    data = [
        ("地质灾害类型", ""),
        ("地质灾害易发程度", "高易发区"),
        ("危险性等级", "小"),
        ("地震基本烈度", "6度"),
        ("地震动峰值加速度", "0.05g"),
        ("防治措施", "工程施工前委托有资质的单位进行详细岩土工程勘察，查明土体的工程地质性质和分布特征，基坑开挖时避免大方量的切坡开挖，合理选择基础持力层，修建好地表排水沟"),
        ("分析结论", "现状条件下，拟建场地无滑坡、崩塌、泥石流等不良地质灾害，地质灾害发育程度弱、危害程度小、危险性小"),
    ]

    for i, (field, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=field)
        ws.cell(row=i, column=2, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 30


def create_social_stability_sheet(wb: Workbook):
    """创建社会稳定Sheet（第4章）"""
    ws = wb.create_sheet("社会稳定")

    # 设置列宽
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50

    # 表头
    headers = ["类别", "项目", "内容"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 数据
    data = [
        # 合法性风险
        ("合法性风险", "风险内容", "项目的决策是否与现行政策、法律、法规相抵触，是否有充分的政策、法律依据"),
        ("合法性风险", "风险等级", "低"),
        ("合法性风险", "防范措施1", "项目经过严格的审查审批和报批程序"),
        ("合法性风险", "防范措施2", "经过严谨科学的可行性研究论证"),
        ("合法性风险", "防范措施3", "建设方案具体、详实，配套措施完善"),
        # 生活环境风险
        ("生活环境风险", "风险内容", "施工期间的扬尘、噪音、建筑垃圾等对周边环境的影响"),
        ("生活环境风险", "风险等级", "低"),
        ("生活环境风险", "防范措施1", "施工单位做好洒水降尘措施"),
        ("生活环境风险", "防范措施2", "合理安排施工时段，避免中午、夜间休息时间作业"),
        ("生活环境风险", "防范措施3", "建筑垃圾清理合理、有序、彻底"),
        ("生活环境风险", "防范措施4", "污水处理厂厂址远离居民区"),
        # 社会环境风险
        ("社会环境风险", "风险内容", "项目建设和运营可能引发的社会稳定问题"),
        ("社会环境风险", "风险等级", "低"),
        ("社会环境风险", "防范措施1", "充分发挥群众来访投诉接待中心功能"),
        ("社会环境风险", "防范措施2", "重视投诉公开电话、网上信访等群众意见反映渠道"),
        ("社会环境风险", "防范措施3", "加强对项目的正面宣传"),
        ("社会环境风险", "防范措施4", "强化利益相关者的参与"),
        ("社会环境风险", "防范措施5", "建立信访稳控工作方案"),
        # 综合结论
        ("综合结论", "社会稳定小结", "本项目社会稳定整体综合风险为低等级，但仍应做好相关的风险防范措施"),
    ]

    for i, (category, item, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=item)
        ws.cell(row=i, column=3, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))
        apply_cell_style(ws.cell(row=i, column=3))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 25


def create_energy_saving_sheet(wb: Workbook):
    """创建节能分析Sheet（第4章）"""
    ws = wb.create_sheet("节能分析")

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60

    # 表头
    ws['A1'] = "项目"
    ws['B1'] = "内容"
    apply_header_style(ws['A1'])
    apply_header_style(ws['B1'])

    # 数据
    data = [
        ("前期工作阶段措施1", "把好选线关，尽可能减少占用土地"),
        ("前期工作阶段措施2", "注重项目方案比选工作，优先选择节约土地的方案"),
        ("前期工作阶段措施3", "细化、优化、深化设计方案比选"),
        ("建设实施阶段措施1", "合理安排施工机械流程，实现效率最大化"),
        ("建设实施阶段措施2", "定期维护检查机械设备，充分发挥机械的生产效率"),
        ("建设实施阶段措施3", "使用节能环保材料"),
        ("施工节能措施1", "管道合理设计管径，减小排水阻力"),
        ("施工节能措施2", "照明光源选用节能、高效的灯具"),
        ("运营节能措施", "选用新型工艺进行污水处理，有效减少用电消耗和资源消耗"),
        ("节能结论", "项目建设过程中积极运用四新技术，采用先进节能技术，有助于节省能源和资源消耗"),
    ]

    for i, (field, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=field)
        ws.cell(row=i, column=2, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 30


def create_rationality_summary_sheet(wb: Workbook):
    """创建合理性小结Sheet（第4章）"""
    ws = wb.create_sheet("合理性小结")

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80

    # 表头
    ws['A1'] = "项目"
    ws['B1'] = "内容"
    apply_header_style(ws['A1'])
    apply_header_style(ws['B1'])

    # 数据
    ws.cell(row=2, column=1, value="合理性结论")
    ws.cell(row=2, column=2, value="综上所述，项目区地质稳定，发生灾害可能性较小，自然条件良好，未压覆现有已探明矿产，对项目区周边环境影响较小，社会稳定性较高，有利于项目的快速开展，同时采用节能技术及方案，有助于节省能源和资源消耗，所以选址是可行的和合理的。")
    apply_cell_style(ws.cell(row=2, column=1))
    apply_cell_style(ws.cell(row=2, column=2))

    for row in range(1, 3):
        ws.row_dimensions[row].height = 60


# ============================================================================
# 第5章：节约集约用地分析 Sheet
# ============================================================================

def create_functional_zone_sheet(wb: Workbook):
    """创建功能分区Sheet（第5章）"""
    ws = wb.create_sheet("功能分区")

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 35

    # 表头
    headers = ["分区名称", "分区面积(平方米)", "占比(%)", "功能描述", "用地依据"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 数据
    data = [
        ("生产区用地", 9208.00, 86.60, "包括二级处理区、深度处理区及污泥处置区", "《城市污水处理工程项目建设标准》（JB198-2022）第十八条"),
        ("生产管理及辅助生产区用地", 1425.00, 13.40, "包括生产管理用房、辅助生产设施等", "《城市污水处理工程项目建设标准》（JB198-2022）第二十七条"),
    ]

    for i, row_data in enumerate(data, start=2):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            apply_cell_style(cell)

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 30


def create_land_scale_sheet(wb: Workbook):
    """创建用地规模Sheet（第5章）"""
    ws = wb.create_sheet("用地规模")

    # 设置列宽
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50

    # 表头
    headers = ["类别", "项目", "内容"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 数据
    data = [
        # 总体指标
        ("总体指标", "项目总用地面积(平方米)", "10633.00"),
        ("总体指标", "建设规模", "6000m³/d"),
        ("总体指标", "标准依据", "《湖北省产业用地目录和用地标准（2023年本）》第六章"),
        ("总体指标", "标准要求范围", "7500～12000平方米"),
        ("总体指标", "是否符合要求", "是"),
        ("总体指标", "对比分析", "本项目拟建设面积为10633.00平方米，满足用地指标要求，符合节约集约用地要求"),
        # 各分区指标
        ("各分区指标", "二级处理区-实际用地面积(平方米)", "6383.00"),
        ("各分区指标", "二级处理区-标准依据", "《城市污水处理工程项目建设标准》（JB198-2022）表2"),
        ("各分区指标", "二级处理区-标准指标值", "9000平方米（按1.5×6000÷10000计算）"),
        ("各分区指标", "二级处理区-是否符合要求", "是"),
        ("各分区指标", "深度处理区-实际用地面积(平方米)", "1216.00"),
        ("各分区指标", "深度处理区-标准依据", "《城市污水处理工程项目建设标准》（JB198-2022）表2"),
        ("各分区指标", "深度处理区-标准指标值", "2700平方米（按0.45×6000÷10000计算）"),
        ("各分区指标", "深度处理区-是否符合要求", "是"),
        ("各分区指标", "污泥处置区-实际用地面积(平方米)", "1609.00"),
        ("各分区指标", "污泥处置区-标准依据", "《城市排水工程规划规范》（GB50318-2017）表1"),
        ("各分区指标", "污泥处置区-标准指标值", "3000平方米"),
        ("各分区指标", "污泥处置区-是否符合要求", "是"),
        # 辅助区用地占比
        ("辅助区用地占比", "实际占比", "13.40%"),
        ("辅助区用地占比", "标准要求范围", "8%～20%"),
        ("辅助区用地占比", "是否符合", "是"),
        # 综合评价
        ("综合评价", "综合评价", "各类用地指标均低于国家标准规定限制，符合节约集约用地要求"),
    ]

    for i, (category, item, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=item)
        ws.cell(row=i, column=3, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))
        apply_cell_style(ws.cell(row=i, column=3))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 25


def create_land_tech_sheet(wb: Workbook):
    """创建节地技术Sheet（第5章）"""
    ws = wb.create_sheet("节地技术")

    # 设置列宽
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 45

    # 表头
    headers = ["类别", "措施名称", "措施描述"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 数据
    data = [
        # 前期工作阶段
        ("前期工作阶段", "把好选线关", "在遵循项目有关设计规范标准的前提下，以尽可能减少占用土地为原则，结合其他影响选线的因素进行深入研究，通过反复比选和论证，确定合理的选址方案"),
        ("前期工作阶段", "注重项目方案比选工作", "项目方案设计中，在深入调查、论证的基础上确定合理的主要控制点，将土地占用情况作为方案选择的重要指标"),
        ("前期工作阶段", "细化、优化、深化设计方案比选", "在满足功能要求、防火要求的前提下，通过合理优化设计可达到节约用地的目的"),
        # 建设实施阶段
        ("建设实施阶段", "施工招标要求", "项目施工招标时，应将耕地保护的有关条款列入招标文件，并严格执行"),
        ("建设实施阶段", "项目法人职责", "项目法人要增强耕地保护意识，统筹工程实施临时用地，加强科学指导"),
        ("建设实施阶段", "施工单位要求", "施工单位要严格控制临时用地数量，施工便道、各种料场、预制场要根据工程进度统筹考虑，尽可能设置在项目用地范围内或利用荒坡、废弃地解决"),
        ("建设实施阶段", "废弃地处理", "建设中废弃的空地要尽可能造地复垦，不能复垦的要尽量绿化，避免闲置浪费"),
        # 综合评价
        ("综合评价", "综合评价", "项目采用了先进的节地技术，符合节约集约用地要求"),
    ]

    for i, (category, item, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=item)
        ws.cell(row=i, column=3, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))
        apply_cell_style(ws.cell(row=i, column=3))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 30


def create_case_comparison_sheet(wb: Workbook):
    """创建案例对比Sheet（第5章）"""
    ws = wb.create_sheet("案例对比")

    # 设置列宽
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 15

    # 表头
    headers = ["案例名称", "案例地点", "建设规模", "用地面积(平方米)", "采用技术", "总投资(万元)"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        apply_header_style(cell)

    # 数据
    data = [
        ("本项目", "", "6000m³/d", 10633.00, "A³/O生化池→二沉池→高效沉淀池+滤布滤池→紫外线消毒", 15938.70),
        ("黄埔镇大雁生活污水处理厂新建工程", "中山市黄圃镇", "30000m³/d", 12367.61, "粗格栅及提升泵房→细格栅及曝气沉砂池→调节池/事故池→A³/O生化池→二沉池→高效沉淀池+滤布滤池→紫外线消毒", 166426.7),
        ("阳春市河西污水处理厂工程", "阳春市河西街道", "10000m³/d", 14800.00, "分体式水解酸化+A2/O+絮凝沉淀+过滤", 25134.2),
    ]

    for i, row_data in enumerate(data, start=2):
        for j, value in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            apply_cell_style(cell)

    # 添加对比结论
    ws.cell(row=5, column=1, value="对比结论")
    ws.cell(row=5, column=2, value="本项目单位用地投资量为1.50万元/平方米，低于对比案例，表明本项目节地水平较先进，节约集约利用效率较高")
    apply_header_style(ws.cell(row=5, column=1))
    ws.merge_cells('B5:F5')
    apply_cell_style(ws.cell(row=5, column=2))

    for row in range(1, 6):
        ws.row_dimensions[row].height = 35


def create_land_use_summary_sheet(wb: Workbook):
    """创建节约集约小结Sheet（第5章）"""
    ws = wb.create_sheet("节约集约小结")

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80

    # 表头
    ws['A1'] = "项目"
    ws['B1'] = "内容"
    apply_header_style(ws['A1'])
    apply_header_style(ws['B1'])

    # 数据
    ws.cell(row=2, column=1, value="节约集约用地小结")
    ws.cell(row=2, column=2, value="项目用地功能分区合理，且各类用地指标均低于国家标准规定限制，同时采用先进节地技术，合理节约利用土地，实现耕地有效保护，与国内同规模同类型项目相比，本项目占地面积较小，投入资金较低，表明本项目节约集约用地方面较为先进，满足项目建设需求。")
    apply_cell_style(ws.cell(row=2, column=1))
    apply_cell_style(ws.cell(row=2, column=2))

    for row in range(1, 3):
        ws.row_dimensions[row].height = 60


# ============================================================================
# 第6章：结论与建议 Sheet
# ============================================================================

def create_conclusion_sheet(wb: Workbook):
    """创建结论建议Sheet（第6章）"""
    ws = wb.create_sheet("结论建议")

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80

    # 表头
    ws['A1'] = "项目"
    ws['B1'] = "内容"
    apply_header_style(ws['A1'])
    apply_header_style(ws['B1'])

    # 数据
    data = [
        # 合法合规性结论
        ("法律法规结论", "符合相关法律法规及政策文件"),
        ("耕地和永久基本农田结论", "不占用耕地和永久基本农田"),
        ("生态保护红线结论", "不占用生态保护红线"),
        ("城镇开发边界结论", "不占用城镇开发边界"),
        ("国土空间规划结论", "符合国土空间总体规划要求"),
        ("综合交通规划符合性", "符合"),
        ("市政基础设施规划符合性", "符合"),
        ("历史文化遗产保护规划符合性", "符合"),
        ("综合防灾工程规划符合性", "符合"),
        ("旅游规划符合性", "符合"),
        # 合理性结论
        ("环境影响结论", "影响较小，有防治措施"),
        ("矿产资源结论", "未压覆已查明矿产"),
        ("地质灾害结论", "受灾程度小，有防治措施"),
        ("社会稳定结论", "影响较小，有相应措施"),
        ("节能结论", "已采用先进节能技术"),
        ("选址合理性综合结论", "选址合理，符合项目建设要求"),
        # 节约集约用地结论
        ("功能分区结论", "功能分区合理"),
        ("用地规模结论", "用地规模各分区合理"),
        ("节地技术结论", "较其他项目节地水平更高"),
        ("节约集约用地综合结论", "符合节约集约用地要求"),
        # 综合论证结论
        ("综合论证结论", "该项目选址符合项目所需的建设场地要求，与各项规划、政策基本协调，项目选址可行"),
        # 建议
        ("建议1", "项目应进一步衔接和协调建设项目选址与各类规划的关系。"),
        ("建议2", "项目选址应进一步分析项目运营对周边敏感目标（如居民区、学校、自然保护区）的潜在环境影响，确保满足卫生防护距离要求，并论证尾水排放对受纳水体的影响是否在区域水环境容量承载范围之内。"),
        ("建议3", "项目选址须精准落入城镇排水与污水处理系统专项规划所确定的设施布局和服务范围内。应重点复核与城镇污水主干管网的衔接可行性，评估提升泵站的设置需求与成本，确保污水能够经济、高效地收集输送。"),
        ("建议4", "项目选址应精确核算厂区防洪排涝标准（通常应高于50年一遇），核实场地标高与周边河流洪水位的关系，确保厂区不受洪涝威胁。对尾水排放口进行水力模型模拟，优化排放方式，减少对河床、岸线的冲刷。"),
        ("建议5", "项目在环评基础上，可针对选址特性，重点深化恶臭气体扩散模拟，精确划定卫生防护距离，并提前规划高效的除臭系统与绿化隔离带布局方案。"),
    ]

    for i, (field, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=field)
        ws.cell(row=i, column=2, value=value)
        apply_cell_style(ws.cell(row=i, column=1))
        apply_cell_style(ws.cell(row=i, column=2))

    for row in range(1, len(data) + 2):
        ws.row_dimensions[row].height = 30


def create_template():
    """创建Excel模板文件"""
    print("开始创建Excel模板...")

    # 确保目录存在
    os.makedirs(TEMPLATE_DIR, exist_ok=True)

    # 创建工作簿
    wb = Workbook()

    # 第1章：项目基本信息
    create_project_info_sheet(wb)

    # 第2章：选址分析
    create_alternatives_sheet(wb)
    create_site_conditions_sheet(wb)
    create_sensitive_sheet(wb)
    create_construction_sheet(wb)
    create_consultation_sheet(wb)
    create_comparison_sheet(wb)

    # 第3章：合法合规性分析
    create_regulation_sheet(wb)
    create_three_lines_sheet(wb)
    create_spatial_planning_sheet(wb)
    create_special_planning_sheet(wb)
    create_other_planning_sheet(wb)
    create_urban_planning_sheet(wb)
    create_compliance_summary_sheet(wb)

    # 第4章：选址合理性分析
    create_environmental_sheet(wb)
    create_mineral_sheet(wb)
    create_geological_sheet(wb)
    create_social_stability_sheet(wb)
    create_energy_saving_sheet(wb)
    create_rationality_summary_sheet(wb)

    # 第5章：节约集约用地分析
    create_functional_zone_sheet(wb)
    create_land_scale_sheet(wb)
    create_land_tech_sheet(wb)
    create_case_comparison_sheet(wb)
    create_land_use_summary_sheet(wb)

    # 第6章：结论与建议
    create_conclusion_sheet(wb)

    # 保存文件
    output_path = os.path.join(TEMPLATE_DIR, "项目数据模板.xlsx")
    wb.save(output_path)
    print(f"模板已保存到: {output_path}")

    return output_path


if __name__ == "__main__":
    template_path = create_template()
    print(f"\n模板创建成功!")
    print(f"路径: {template_path}")
    print(f"\n模板包含以下Sheet:")
    print("第1章: 项目基本信息")
    print("第2章: 备选方案、场地条件、敏感条件、施工运营、征求意见、方案比选")
    print("第3章: 法规政策、三线分析、国土空间规划、专项规划、其他规划、城乡总体规划、合法合规小结")
    print("第4章: 环境影响、矿产资源、地质灾害、社会稳定、节能分析、合理性小结")
    print("第5章: 功能分区、用地规模、节地技术、案例对比、节约集约小结")
    print("第6章: 结论建议")
    print("\n用户可以复制此模板并填写数据，然后使用ExcelParser解析。")