"""
扩展Excel模板 - 添加第3章合法合规性分析数据Sheet

根据sample.md中第3章的内容，添加以下Sheet：
1. 法规政策 - 法规政策符合性数据
2. 三线分析 - 三线协调分析数据
3. 国土空间规划 - 国土空间规划符合性数据
4. 专项规划 - 专项规划符合性数据
5. 其他规划 - 其他相关规划符合性数据
6. 城乡总体规划 - 过渡期城乡总体规划数据
"""

import os
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 添加项目根目录到Python路径
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)


def add_chapter3_sheets():
    """添加第3章数据Sheet到Excel模板"""
    
    # 文件路径
    excel_path = os.path.join(
        project_root,
        "templates",
        "excel_templates",
        "项目数据模板.xlsx"
    )
    
    output_path = os.path.join(
        project_root,
        "templates",
        "excel_templates",
        "项目数据模板_第3章.xlsx"
    )
    
    print("=" * 80)
    print("扩展Excel模板 - 添加第3章数据Sheet")
    print("=" * 80)
    print(f"输入文件: {excel_path}")
    print(f"输出文件: {output_path}")
    print()
    
    if not os.path.exists(excel_path):
        print(f"✗ 源文件不存在: {excel_path}")
        return False
    
    # 加载工作簿
    wb = load_workbook(excel_path)
    print(f"✓ 已加载工作簿，包含 {len(wb.sheetnames)} 个Sheet:")
    for name in wb.sheetnames:
        print(f"  - {name}")
    print()
    
    # =========================================================================
    # Sheet 1: 法规政策
    # =========================================================================
    print("创建Sheet: 法规政策...")
    ws = wb.create_sheet("法规政策", 7)  # 在第7个位置插入
    
    # 表头样式
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # 写入表头
    headers = ["法规名称", "发布单位", "发布时间", "符合性分析", "符合性结论"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 写入示例数据
    sample_data = [
        ["《产业结构调整指导目录（2024年本）》", "国家发展和改革委员会", "2024年", "项目属于第一类鼓励类：四十二、环境保护与资源节约综合利用中的3.城镇污水垃圾处理，符合国家产业政策。", "符合"],
        ["《划拨用地目录》", "自然资源部", "2001年", "项目符合城市基础设施用地和公益事业用地（三）城市基础设施用地5.环境卫生设施：包括雨水处理设施、污水处理厂，符合供地政策。", "符合"],
        ["《城镇污水处理厂污染物排放标准》（GB18918-2002）", "国家环境保护总局", "2002年", "出水水质满足一级A标准。", "符合"],
        ["《城市污水处理工程项目建设标准》（JB198-2022）", "住房和城乡建设部", "2022年", "项目用地规模符合建设标准要求。", "符合"],
    ]
    
    for row in sample_data:
        ws.append(row)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 12
    
    print(f"  ✓ 已添加 {len(sample_data)} 条示例数据")
    
    # =========================================================================
    # Sheet 2: 三线分析
    # =========================================================================
    print("创建Sheet: 三线分析...")
    ws = wb.create_sheet("三线分析", 8)
    
    headers = ["项目", "内容"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    sample_data = [
        ["是否占用耕地", "否"],
        ["占用耕地面积（平方米）", "0"],
        ["是否占用永久基本农田", "否"],
        ["占用永久基本农田面积（平方米）", "0"],
        ["是否占用生态保护红线", "否"],
        ["占用生态保护红线面积（平方米）", "0"],
        ["是否位于城镇开发边界内", "否"],
        ["城镇开发边界说明", "项目位于城镇开发边界外"],
        ["符合性说明", "项目属于《省自然资源厅关于加强'三区三线'实施管理的意见》（鄂自然资函〔2023〕385号）文件规定的单独选址项目用地清单中民生基础设施项目，符合单独选址要求。拟选址项目区不涉及压占耕地、不涉及压占永久基本农田，不涉及压占生态保护红线，不位于城镇开发边界内。符合永久基本农田管控规则，符合生态保护红线管控要求。"],
        ["数据来源", "2023年国土变更调查数据"],
    ]
    
    for row in sample_data:
        ws.append(row)
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 80
    
    print(f"  ✓ 已添加 {len(sample_data)} 条示例数据")
    
    # =========================================================================
    # Sheet 3: 国土空间规划
    # =========================================================================
    print("创建Sheet: 国土空间规划...")
    ws = wb.create_sheet("国土空间规划", 9)
    
    headers = ["类别", "项目", "内容"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    sample_data = [
        ["一张图落位", "是否上图落位", "是"],
        ["一张图落位", "重点项目库名称", "生态修复重点工程"],
        ["一张图落位", "项目类型", "生态修复类项目"],
        ["一张图落位", "落位说明", "项目选址方案符合兴山县国土空间总体规划，本项目在兴山县国土空间规划'一张图'上，已列入规划生态修复重点工程。"],
        ["功能分区准入", "城镇建设适宜性", "城镇建设适宜区"],
        ["功能分区准入", "生态保护重要性", "不属于生态保护极重要区"],
        ["功能分区准入", "农业生产适宜性", "农业生产不适宜区"],
        ["功能分区准入", "符合性说明", "项目选址方案在城镇适宜性评价结果图中属于城镇建设适宜区，在生态保护重要性评价结果图中不属于生态保护极重要区，于农业生产适宜性评价结果图中属于农业生产不适宜区，完全符合生态重要性等级、农业生产适宜等级和建设开发适宜等级等方面的符合性。"],
        ["用途管制", "符合性说明", "项目符合国土空间用途管制相关要求，用地性质为排水用地，符合规划用途。"],
        ["总体格局", "符合性说明", "项目整体符合国土空间总体规划的布局要求，已列入规划生态修复重点工程。"],
        ["总体结论", "总体符合性结论", "项目与国土空间总体规划符合性较高。"],
    ]
    
    for row in sample_data:
        ws.append(row)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 80
    
    print(f"  ✓ 已添加 {len(sample_data)} 条示例数据")
    
    # =========================================================================
    # Sheet 4: 专项规划
    # =========================================================================
    print("创建Sheet: 专项规划...")
    ws = wb.create_sheet("专项规划", 10)
    
    headers = ["规划类型", "规划名称", "符合性分析", "符合性结论"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    sample_data = [
        ["综合交通规划", "兴山县国土空间总体规划（2021-2035年）-综合交通规划", "项目拟选址位于规划'四纵三横两支'的公路骨架网络的'横三'312省道沿线，该项目未占用S312省道沿线，并沿S312省道沿线一侧预留绿化带。", "符合"],
        ["市政基础设施规划", "兴山县国土空间总体规划（2021-2035年）-市政基础设施规划", "项目拟选址与市政基础设施规划中能源供应体系、通讯网络体系、水循环利用与污水治理体系相关规划均不冲突。", "符合"],
        ["历史文化遗产保护规划", "兴山县国土空间总体规划（2021-2035年）-历史文化遗产保护规划", "项目所在区域无压占历史文化名城、名镇、名村保护范围等，均未发现古城镇、古建筑、古村落，文物古迹保护等保护对象。", "符合"],
        ["综合防灾工程规划", "兴山县国土空间总体规划（2021-2035年）-综合防灾减灾体系", "项目所在区域位于一般防治区，主要面临轻度地质灾害和气象灾害风险，不与防灾重点工程冲突。", "符合"],
        ["旅游规划", "兴山县国土空间总体规划（2021-2035年）-城镇空间布局结构图", "项目所在区域位于峡口旅游物流组团片区内部，周边为香溪河绿色山水风景廊道轴线，不与现有旅游规划项目冲突，项目建成有助于维护周边环境。", "符合"],
        ["环境保护规划", "宜昌市环境保护总体规划（2013-2030年）", "项目拟选址位于生态黄线区，主要为小型点状开发的污水处理项目，针对生活污水处理加强了生态治理和修复。", "符合"],
        ["自然保护地规划", "湖北三峡万朝山省级自然保护区", "项目用地不在湖北三峡万朝山省级自然保护区范围内。", "符合"],
    ]
    
    for row in sample_data:
        ws.append(row)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    
    print(f"  ✓ 已添加 {len(sample_data)} 条示例数据")
    
    # =========================================================================
    # Sheet 5: 其他规划
    # =========================================================================
    print("创建Sheet: 其他规划...")
    ws = wb.create_sheet("其他规划", 11)
    
    headers = ["规划类型", "规划名称", "符合性分析", "符合性结论"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    sample_data = [
        ["国民经济和社会发展规划", "《兴山县国民经济和社会发展第十四个五年规划和2035年远景目标纲要》", "项目可对香溪河左岸峡口片区的生活污水进行系统治理和循环利用，与《规划纲要》中'持续抓好香溪河流域生态保护和修复'要点符合。", "符合"],
        ["生态环境保护规划", "《宜昌市环境保护总体规划（2013-2030年）》", "项目拟选址位于生态黄线区，符合管控要求。", "符合"],
        ["三线一单生态环境分区管控", "《宜昌市'三线一单'生态环境分区管控实施方案》", "项目拟选址位于重点管控单元，编号为ZH42052620003，项目不属于该单元空间布局约束禁止项目，符合管控单元污染物排放和环境风险防控相关要求。", "符合"],
        ["综合交通体系规划", "《宜昌市综合交通体系规划（2011-2030年）》", "项目拟选址位于规划的二级公路S312省道沿线，该项目未占用S312省道沿线，并沿S312省道沿线一侧预留绿化带。", "符合"],
    ]
    
    for row in sample_data:
        ws.append(row)
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    
    print(f"  ✓ 已添加 {len(sample_data)} 条示例数据")
    
    # =========================================================================
    # Sheet 6: 城乡总体规划
    # =========================================================================
    print("创建Sheet: 城乡总体规划...")
    ws = wb.create_sheet("城乡总体规划", 12)
    
    headers = ["项目", "内容"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    sample_data = [
        ["规划名称", "《宜昌市兴山县峡口镇总体规划（2014-2030）》"],
        ["规划期限", "2014-2030年"],
        ["空间管制分区", "适建区"],
        ["符合性分析", "项目拟选址位于镇域空间管制分类中适建区，适建区主要为工程地质条件良好、地势相对平坦、没有其它建设限制条件的区域，是城市发展优先选择的地区，需根据资源环境条件，合理确定开发模式、建设规模和开发强度。本项目建设规模、开发强度均较小，符合所在地经批复的城市总体规划布局的要求。"],
        ["符合性结论", "符合"],
    ]
    
    for row in sample_data:
        ws.append(row)
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80
    
    print(f"  ✓ 已添加 {len(sample_data)} 条示例数据")
    
    # =========================================================================
    # Sheet 7: 合法合规小结
    # =========================================================================
    print("创建Sheet: 合法合规小结...")
    ws = wb.create_sheet("合法合规小结", 13)
    
    headers = ["项目", "内容"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    sample_data = [
        ["合法合规小结", "综合前述项目与相关法律法规、政策文件的符合性分析、与'三线'和耕地等各类空间的协调分析、与国土空间总体规划的符合性分析、与专项规划的符合性分析以及与其他相关规划的符合性分析，建设项目总体上属于合法合规。"],
    ]
    
    for row in sample_data:
        ws.append(row)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 80
    
    print(f"  ✓ 已添加 {len(sample_data)} 条示例数据")
    
    # 保存文件
    print()
    print("保存文件...")
    wb.save(output_path)
    wb.close()
    
    print(f"✓ 已保存到: {output_path}")
    print()
    print("=" * 80)
    print("Excel模板扩展完成！")
    print("=" * 80)
    print()
    print("新增Sheet列表:")
    print("  1. 法规政策 - 法规政策符合性数据（4条）")
    print("  2. 三线分析 - 三线协调分析数据（10条）")
    print("  3. 国土空间规划 - 国土空间规划符合性数据（11条）")
    print("  4. 专项规划 - 专项规划符合性数据（7条）")
    print("  5. 其他规划 - 其他相关规划符合性数据（4条）")
    print("  6. 城乡总体规划 - 过渡期城乡总体规划数据（5条）")
    print("  7. 合法合规小结 - 合法合规性分析小结（1条）")
    print()
    print("下一步：更新ExcelParser添加解析逻辑")
    
    return True


if __name__ == "__main__":
    success = add_chapter3_sheets()
    sys.exit(0 if success else 1)