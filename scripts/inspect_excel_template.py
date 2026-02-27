"""
查看Excel模板的Sheet结构
"""

import os
import sys

# 添加项目根目录到Python路径
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

from openpyxl import load_workbook

def inspect_excel_template():
    """查看Excel模板的结构"""
    excel_path = os.path.join(
        project_root,
        "templates",
        "excel_templates",
        "项目数据模板.xlsx"
    )
    
    print("=" * 80)
    print("Excel模板结构分析")
    print("=" * 80)
    print(f"文件路径: {excel_path}")
    print()
    
    if not os.path.exists(excel_path):
        print(f"✗ 文件不存在: {excel_path}")
        return
    
    try:
        # 加载工作簿
        wb = load_workbook(excel_path, data_only=True)
        
        print(f"工作簿包含 {len(wb.sheetnames)} 个Sheet:")
        print()
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print("-" * 80)
            print(f"Sheet名称: {sheet_name}")
            print("-" * 80)
            
            # 获取数据区域
            data = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row_idx > 20:  # 只显示前20行
                    print("  ...")
                    break
                
                # 跳过完全空的行
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue
                
                # 格式化显示
                display_row = []
                for cell in row:
                    if cell is None:
                        display_row.append("")
                    else:
                        display_row.append(str(cell)[:20])  # 截断到20字符
                data.append(display_row)
                
                if data:
                    print(f"  Row {row_idx}: {' | '.join(display_row)}")
            
            print()
        
        wb.close()
        
        print("=" * 80)
        print("分析完成")
        print("=" * 80)
        
    except Exception as e:
        print(f"✗ 读取Excel文件失败: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    inspect_excel_template()