"""
Excel数据解析服务

负责读取用户填写的Excel模板，解析为Pydantic数据模型。
支持第1章项目概况和第2章选址分析数据的解析。
"""

import os
from typing import Dict, Any, List, Optional, Tuple
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from src.models.project_overview_data import ProjectOverviewData
from src.models.site_selection_data import (
    SiteSelectionData,
    SiteAlternative,
    SiteNaturalConditions,
    SiteExternalConditions,
    SiteSensitiveConditions,
    ConstructionConditions,
    PlanningImpact,
    ConsultationOpinion,
    SchemeComparison,
)
from src.services.data_validator import DataValidator, ValidationReport



try:
    from src.models.compliance_data import ComplianceData
except ImportError:
    ComplianceData = None

try:
    from src.models.rationality_data import (
        RationalityData,
        EnvironmentalImpactAnalysis,
        AtmosphericImpact,
        NoiseImpact,
        WaterImpact,
        SolidWasteImpact,
        TrafficImpact,
        EcologicalRestoration,
        MineralResourceAnalysis,
        GeologicalHazardAnalysis,
        SocialStabilityAnalysis,
        LegalityRiskAnalysis,
        LivingEnvironmentRisk,  # 修正类名
        SocialEnvironmentRisk,   # 修正类名
        EnergySavingAnalysis,
    )
except ImportError:
    RationalityData = None
    EnvironmentalImpactAnalysis = None
    AtmosphericImpact = None
    NoiseImpact = None
    WaterImpact = None
    SolidWasteImpact = None
    TrafficImpact = None
    EcologicalRestoration = None
    MineralResourceAnalysis = None
    GeologicalHazardAnalysis = None
    SocialStabilityAnalysis = None
    LegalityRiskAnalysis = None
    LivingEnvironmentRisk = None
    SocialEnvironmentRisk = None
    EnergySavingAnalysis = None

try:
    from src.models.land_use_data import (
        LandUseData,
        FunctionalZone,
        OverallLandIndicator,
        SubAreaIndicator,
        LandScaleAnalysis,
        PreConstructionMeasure,
        ConstructionPhaseMeasure,
        LandSavingTechnology,
        ComparisonCase,
        CaseComparison,
    )
except ImportError:
    LandUseData = None
    FunctionalZone = None
    OverallLandIndicator = None
    SubAreaIndicator = None
    LandScaleAnalysis = None
    PreConstructionMeasure = None
    ConstructionPhaseMeasure = None
    LandSavingTechnology = None
    ComparisonCase = None
    CaseComparison = None

try:
    from src.models.conclusion_data import (
        ConclusionData,
        合规性结论,
        合理性结论,
        节约集约结论,
        建议项,
    )
except ImportError:
    ConclusionData = None
    合规性结论 = None
    合理性结论 = None
    节约集约结论 = None
    建议项 = None

from src.utils.logger import logger



class ExcelParseError(Exception):
    """Excel解析错误"""
    pass


class ExcelParser:
    """
    Excel数据解析器

    解析项目数据Excel文件，支持以下Sheet:
    - 项目基本信息：第 1 章数据
    - 备选方案：第 2 章备选方案数据
    - 场地条件：第 2 章自然条件和外部条件
    - 敏感条件：第 2 章敏感条件
    - 施工运营：第 2 章施工运营条件
    - 征求意见：第 2 章征求意见情况
    - 方案比选：第 2 章方案比选数据
    - 法规政策：第 3 章法规政策符合性
    - 三线分析：第 3 章三线协调分析
    - 国土空间规划：第 3 章国土空间规划符合性
    - 专项规划：第 3 章专项规划符合性
    - 其他规划：第 3 章其他规划符合性
    - 城乡总体规划：第 3 章城乡总体规划符合性
    - 合法合规小结：第 3 章小结
    - 合理性小结：第 4 章选址合理性分析小结
    - 功能分区：第 5 章功能分区情况
    - 用地规模：第 5 章用地规模合理性分析
    - 节地技术：第 5 章节地技术措施
    - 案例对比：第 5 章案例对比分析
    - 节约集约小结：第 5 章节约集约用地分析小结
    - 节约集约小结：第 5 章节约集约用地分析小结
    - 结论建议：第 6 章结论与建议数据
    """
    # Sheet名称常量
    SHEET_PROJECT_INFO = "项目基本信息"
    SHEET_ALTERNATIVES = "备选方案"
    SHEET_SITE_CONDITIONS = "场地条件"
    SHEET_SENSITIVE = "敏感条件"
    SHEET_CONSTRUCTION = "施工运营"
    SHEET_COMPARISON = "方案比选"
    SHEET_CONSULTATION = "征求意见"
    
    # 第 3 章 Sheet 常量
    SHEET_REGULATION = "法规政策"
    SHEET_THREE_LINES = "三线分析"
    SHEET_SPATIAL_PLANNING = "国土空间规划"
    SHEET_SPECIAL_PLANNING = "专项规划"
    SHEET_OTHER_PLANNING = "其他规划"
    SHEET_URBAN_PLANNING = "城乡总体规划"
    SHEET_COMPLIANCE_SUMMARY = "合法合规小结"
    
    # 第 4 章 Sheet 常量
    SHEET_ENVIRONMENTAL = "环境影响"
    SHEET_MINERAL = "矿产资源"
    SHEET_GEOLOGICAL = "地质灾害"
    SHEET_SOCIAL_STABILITY = "社会稳定"
    SHEET_ENERGY_SAVING = "节能分析"
    SHEET_RATIONALITY_SUMMARY = "合理性小结"
    
    # 第 5 章 Sheet 常量
    SHEET_FUNCTIONAL_ZONE = "功能分区"
    SHEET_LAND_SCALE = "用地规模"
    SHEET_LAND_TECH = "节地技术"
    SHEET_CASE_COMPARISON = "案例对比"
    SHEET_LAND_USE_SUMMARY = "节约集约小结"
    
    # 第 6 章 Sheet 常量
    SHEET_CONCLUSION = "结论建议"
    
    def __init__(self, file_path: str):
        """
        初始化解析器

        Args:
            file_path: Excel文件路径
        """
        self.file_path = file_path
        self.workbook: Optional[Workbook] = None
        self._validate_file()

    def _validate_file(self):
        """验证文件是否存在且为Excel格式"""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel文件不存在: {self.file_path}")

        if not self.file_path.lower().endswith(('.xlsx', '.xlsm')):
            raise ExcelParseError(f"不支持的文件格式: {self.file_path}")

    def _load_workbook(self):
        """加载Excel工作簿"""
        if self.workbook is None:
            logger.info(f"加载Excel文件: {self.file_path}")
            self.workbook = load_workbook(self.file_path, data_only=True)
            logger.info(f"工作簿包含Sheet: {self.workbook.sheetnames}")

    def _get_sheet(self, sheet_name: str) -> Optional[Worksheet]:
        """
        获取指定名称的Sheet

        Args:
            sheet_name: Sheet名称

        Returns:
            Worksheet对象，如果不存在则返回None
        """
        self._load_workbook()
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        return None

    def _read_key_value_sheet(self, sheet: Worksheet) -> Dict[str, str]:
        """
        读取键值对格式的Sheet

        格式: 第一列为字段名，第二列为字段值

        Args:
            sheet: Worksheet对象

        Returns:
            键值对字典
        """
        result = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                key = str(row[0]).strip()
                value = str(row[1]).strip() if row[1] is not None else ""
                result[key] = value
        return result

    def _read_category_sheet(self, sheet: Worksheet) -> Dict[str, Dict[str, Any]]:
        """
        读取分类格式的Sheet

        格式: 第一列类别，第二列项目，第三列内容

        Args:
            sheet: Worksheet对象

        Returns:
            分类字典
        """
        result: Dict[str, Dict[str, Any]] = {}
        current_category = None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                # 新类别
                current_category = str(row[0]).strip()
                if current_category not in result:
                    result[current_category] = {}

            if current_category and row[1] is not None:
                item_name = str(row[1]).strip()
                item_value = row[2] if len(row) > 2 and row[2] is not None else ""
                result[current_category][item_name] = item_value

        return result

    def _read_table_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """
        读取表格格式的Sheet

        格式: 第一行为表头，后续行为数据

        Args:
            sheet: Worksheet对象

        Returns:
            数据行列表
        """
        result = []
        headers = []

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                # 第一行为表头
                headers = [str(cell).strip() if cell else "" for cell in row]
            else:
                # 数据行
                row_data = {}
                for j, cell in enumerate(row):
                    if j < len(headers) and headers[j]:
                        row_data[headers[j]] = cell
                if row_data:
                    result.append(row_data)

        return result

    def parse_project_overview(self) -> ProjectOverviewData:
        """
        解析第1章：项目基本信息

        Returns:
            ProjectOverviewData对象
        """
        logger.info("解析项目基本信息...")

        sheet = self._get_sheet(self.SHEET_PROJECT_INFO)
        if sheet is None:
            raise ExcelParseError(f"缺少Sheet: {self.SHEET_PROJECT_INFO}")

        data = self._read_key_value_sheet(sheet)
        logger.info(f"读取到 {len(data)} 个项目信息字段")

        try:
            project_data = ProjectOverviewData(
                项目名称=data.get("项目名称", ""),
                项目代码=data.get("项目代码"),
                建设单位=data.get("建设单位", ""),
                建设性质=data.get("建设性质", ""),
                项目投资=data.get("项目投资", ""),
                项目选址=data.get("项目选址", data.get("项目位置", "")),
                建设内容=data.get("建设内容", ""),
                建设规模=data.get("建设规模"),
                建设期限=data.get("建设期限"),
            )
            logger.info(f"项目基本信息解析成功: {project_data.项目名称}")
            return project_data

        except ValidationError as e:
            raise ExcelParseError(f"项目基本信息数据验证失败: {str(e)}")

    def parse_site_selection(self) -> SiteSelectionData:
        """
        解析第2章：选址分析数据

        Returns:
            SiteSelectionData对象
        """
        logger.info("解析选址分析数据...")

        # 解析各个部分
        project_info = self._parse_project_info_for_site()
        alternatives = self._parse_alternatives()
        site_conditions = self._parse_site_conditions()
        sensitive_conditions = self._parse_sensitive_conditions()
        construction_conditions = self._parse_construction_conditions()
        planning_impact = self._parse_planning_impact()
        consultation_opinions = self._parse_consultation_opinions()
        comparison = self._parse_scheme_comparison()
        principles = self._parse_site_principles()

        try:
            site_data = SiteSelectionData(
                项目基本信息=project_info,
                选址原则=principles,
                备选方案=alternatives,
                场地自然条件=site_conditions["自然条件"],
                外部配套条件=site_conditions["外部配套"],
                选址敏感条件=sensitive_conditions,
                施工运营条件=construction_conditions,
                规划影响=planning_impact,
                征求意见情况=consultation_opinions,
                方案比选=comparison,
            )
            logger.info("选址分析数据解析成功")
            return site_data

        except ValidationError as e:
            raise ExcelParseError(f"选址分析数据验证失败: {str(e)}")

    def _parse_project_info_for_site(self) -> Dict[str, str]:
        """解析项目基本信息（用于选址数据）"""
        sheet = self._get_sheet(self.SHEET_PROJECT_INFO)
        if sheet:
            return self._read_key_value_sheet(sheet)
        return {}

    def _parse_alternatives(self) -> List[SiteAlternative]:
        """解析备选方案"""
        logger.info("  解析备选方案...")
        sheet = self._get_sheet(self.SHEET_ALTERNATIVES)

        if sheet is None:
            raise ExcelParseError(f"缺少Sheet: {self.SHEET_ALTERNATIVES}")

        rows = self._read_table_sheet(sheet)
        alternatives = []

        for row in rows:
            try:
                # 获取面积值（支持多种列名格式）
                面积值 = row.get("面积") or row.get("面积(平方米)") or row.get("面积（平方米）") or 0
                if 面积值:
                    面积值 = float(面积值) if 面积值 else 0

                # 构建四至范围
                四至范围 = {
                    "东": row.get("东", row.get("东侧", "")),
                    "南": row.get("南", row.get("南侧", "")),
                    "西": row.get("西", row.get("西侧", "")),
                    "北": row.get("北", row.get("北侧", "")),
                }

                # 构建土地利用现状
                土地利用现状 = {}
                for key in ["农村道路", "林地", "园地", "耕地", "建设用地", "交通运输用地", "农村宅基地"]:
                    if key in row and row[key]:
                        土地利用现状[key] = str(row[key])

                alt = SiteAlternative(
                    方案编号=str(row.get("方案编号", "")),
                    方案名称=row.get("方案名称", ""),
                    位置=row.get("位置", ""),
                    面积=面积值,
                    四至范围=四至范围,
                    土地利用现状=土地利用现状 if 土地利用现状 else {"未分类": "0"},
                    是否占用耕地=self._parse_bool(row.get("是否占用耕地", "否")),
                    是否占用永久基本农田=self._parse_bool(row.get("是否占用永久基本农田", "否")),
                    是否涉及未利用地=self._parse_bool(row.get("是否涉及未利用地", "否")),
                    建设内容=row.get("建设内容", ""),
                    工艺流程=row.get("工艺流程"),
                    出水标准=row.get("出水标准"),
                )
                alternatives.append(alt)

            except Exception as e:
                logger.warning(f"解析备选方案行失败: {str(e)}")
                continue

        if len(alternatives) < 2:
            raise ExcelParseError(f"至少需要2个备选方案，当前解析到{len(alternatives)}个")

        logger.info(f"  解析到 {len(alternatives)} 个备选方案")
        return alternatives

    def _parse_site_conditions(self) -> Dict[str, Any]:
        """解析场地条件"""
        logger.info("  解析场地条件...")
        sheet = self._get_sheet(self.SHEET_SITE_CONDITIONS)

        result = {
            "自然条件": None,
            "外部配套": None,
        }

        if sheet is None:
            raise ExcelParseError(f"缺少Sheet: {self.SHEET_SITE_CONDITIONS}")

        categories = self._read_category_sheet(sheet)

        # 构建自然条件
        natural_conditions = SiteNaturalConditions(
            地形地貌=categories.get("地形地貌", {}),
            气候=categories.get("气候", {}),
            区域地质构造=categories.get("区域地质构造", categories.get("地质构造", {})),
            水文地质条件=categories.get("水文地质条件", categories.get("水文地质", {})),
            工程地质=categories.get("工程地质", {}),
            地震=categories.get("地震", {}),
        )
        result["自然条件"] = natural_conditions

        # 构建外部配套条件
        external = categories.get("外部配套条件", categories.get("外部配套", {}))
        result["外部配套"] = SiteExternalConditions(
            周边建筑物=external.get("周边建筑物", ""),
            供水=external.get("供水", ""),
            供电=external.get("供电", ""),
            通讯=external.get("通讯", ""),
            交通=external.get("交通", ""),
            建材来源=external.get("建材来源", ""),
            是否压覆文物=self._parse_bool(external.get("是否压覆文物", "否")),
            是否影响防洪=self._parse_bool(external.get("是否影响防洪", "否")),
        )

        logger.info("  场地条件解析成功")
        return result

    def _parse_sensitive_conditions(self) -> SiteSensitiveConditions:
        """解析敏感条件"""
        logger.info("  解析敏感条件...")
        sheet = self._get_sheet(self.SHEET_SENSITIVE)

        if sheet is None:
            logger.warning(f"缺少Sheet: {self.SHEET_SENSITIVE}，使用默认值")
            return SiteSensitiveConditions(
                历史保护={},
                生态保护={},
                矿产资源={},
                安全防护={},
                重要设施={},
                耕地和基本农田={},
                生态保护红线={},
            )

        categories = self._read_category_sheet(sheet)

        # 处理布尔值
        def to_bool_dict(d: Dict) -> Dict[str, bool]:
            return {k: self._parse_bool(v) for k, v in d.items()}

        def to_str_dict(d: Dict) -> Dict[str, str]:
            return {k: str(v) if v is not None else "" for k, v in d.items()}

        return SiteSensitiveConditions(
            历史保护=to_bool_dict(categories.get("历史保护", {})),
            生态保护=to_str_dict(categories.get("生态保护", {})),
            矿产资源=to_bool_dict(categories.get("矿产资源", {})),
            安全防护=to_bool_dict(categories.get("安全防护", {})),
            重要设施=to_str_dict(categories.get("重要设施", {})),
            耕地和基本农田=to_bool_dict(categories.get("耕地和基本农田", {})),
            生态保护红线=to_bool_dict(categories.get("生态保护红线", {})),
        )

    def _parse_construction_conditions(self) -> ConstructionConditions:
        """解析施工运营条件"""
        logger.info("  解析施工运营条件...")
        sheet = self._get_sheet(self.SHEET_CONSTRUCTION)

        if sheet is None:
            raise ExcelParseError(f"缺少Sheet: {self.SHEET_CONSTRUCTION}")

        data = self._read_key_value_sheet(sheet)

        return ConstructionConditions(
            方案一总投资=data.get("方案一总投资", ""),
            方案二总投资=data.get("方案二总投资", ""),
            政府支持=data.get("政府支持", ""),
            群众支持=data.get("群众支持", ""),
            征地拆迁=data.get("征地拆迁"),
            施工难度=data.get("施工难度", ""),
            材料供应=data.get("材料供应", ""),
        )

    def _parse_planning_impact(self) -> PlanningImpact:
        """解析规划影响"""
        logger.info("  解析规划影响...")
        sheet = self._get_sheet(self.SHEET_CONSTRUCTION)

        if sheet is None:
            # 使用默认值
            return PlanningImpact(
                是否符合国土空间总体规划=True,
                是否列入重点项目库=True,
                重点项目库名称="",
                对区域发展作用="",
            )

        data = self._read_key_value_sheet(sheet)

        # 尝试从施工运营Sheet读取规划影响相关字段
        return PlanningImpact(
            是否符合国土空间总体规划=self._parse_bool(data.get("是否符合国土空间总体规划", "是")),
            是否列入重点项目库=self._parse_bool(data.get("是否列入重点项目库", "是")),
            重点项目库名称=data.get("重点项目库名称"),
            对区域发展作用=data.get("对区域发展作用", ""),
        )

    def _parse_consultation_opinions(self) -> List[ConsultationOpinion]:
        """解析征求意见情况"""
        logger.info("  解析征求意见情况...")
        sheet = self._get_sheet(self.SHEET_CONSULTATION)

        if sheet is None:
            raise ExcelParseError(f"缺少Sheet: {self.SHEET_CONSULTATION}")

        rows = self._read_table_sheet(sheet)
        opinions = []

        for row in rows:
            try:
                opinion = ConsultationOpinion(
                    部门=row.get("部门", ""),
                    日期=str(row.get("日期", "")),
                    复函标题=row.get("复函标题", ""),
                    结论=row.get("结论", ""),
                )
                opinions.append(opinion)
            except Exception as e:
                logger.warning(f"解析征求意见行失败: {str(e)}")
                continue

        if len(opinions) < 3:
            raise ExcelParseError(f"至少需要3个部门意见，当前解析到{len(opinions)}个")

        logger.info(f"  解析到 {len(opinions)} 条征求意见")
        return opinions

    def _parse_scheme_comparison(self) -> SchemeComparison:
        """解析方案比选"""
        logger.info("  解析方案比选...")
        sheet = self._get_sheet(self.SHEET_COMPARISON)

        if sheet is None:
            raise ExcelParseError(f"缺少Sheet: {self.SHEET_COMPARISON}")

        data = self._read_key_value_sheet(sheet)

        # 解析比选因子（可能是逗号分隔的字符串或列表）
        factors_str = data.get("比选因子", "")
        if isinstance(factors_str, str):
            factors = [f.strip() for f in factors_str.split(",") if f.strip()]
        else:
            factors = []

        # 解析推荐理由
        reasons_str = data.get("推荐理由", "")
        if isinstance(reasons_str, str):
            reasons = [r.strip() for r in reasons_str.split(",") if r.strip()]
        else:
            reasons = []

        return SchemeComparison(
            比选因子=factors if factors else ["场地自然条件", "外部配套条件", "选址敏感条件", "施工运营条件", "规划影响条件"],
            推荐方案=data.get("推荐方案", "方案一"),
            推荐理由=reasons if reasons else ["投资较低"],
        )

    def _parse_site_principles(self) -> List[str]:
        """解析选址原则"""
        sheet = self._get_sheet(self.SHEET_PROJECT_INFO)

        if sheet is None:
            return [
                "符合规划要求",
                "不占优质耕地",
                "尽量不迁移民",
                "避免敏感区域",
                "基础设施优先",
                "集约节约利用",
                "方便施工运营",
                "安全可靠",
            ]

        data = self._read_key_value_sheet(sheet)
        principles_str = data.get("选址原则", "")

        if isinstance(principles_str, str) and principles_str:
            return [p.strip() for p in principles_str.split(",") if p.strip()]

        return [
            "符合规划要求",
            "不占优质耕地",
            "尽量不迁移民",
            "避免敏感区域",
            "基础设施优先",
        ]

    def _parse_bool(self, value: Any) -> bool:
        """
        解析布尔值

        Args:
            value: 输入值

        Returns:
            布尔值
        """
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ("是", "true", "yes", "1", "√")
        if isinstance(value, (int, float)):
            return bool(value)
        return False

    def parse_compliance(self) -> 'ComplianceData':
        """
        解析合法合规性分析数据（第3章）
        
        Returns:
            ComplianceData: 合法合规性分析数据模型
        """
        if ComplianceData is None:
            raise ImportError("ComplianceData模型未导入，请检查compliance_data模块")
        
        logger.info("开始解析合法合规性分析数据...")
        
        project_basic = self.parse_project_overview().to_dict()
        
        # 解析各Sheet
        regulation = self._parse_regulation()
        three_lines = self._parse_three_lines()
        spatial_planning = self._parse_spatial_planning()
        special_planning = self._parse_special_planning()
        other_planning = self._parse_other_planning()
        urban_planning = self._parse_urban_planning()
        summary = self._parse_compliance_summary()
        
        logger.info("合法合规性分析数据解析完成")
        
        return ComplianceData(
            项目基本信息=project_basic,
            产业政策符合性=regulation[0] if len(regulation) > 0 else None,
            供地政策符合性=regulation[1] if len(regulation) > 1 else None,
            其他法规符合性=regulation[2:] if len(regulation) > 2 else None,
            三线协调分析=three_lines,
            国土空间规划符合性=spatial_planning,
            专项规划符合性=special_planning,
            其他规划符合性=other_planning,
            城乡总体规划符合性=urban_planning,
            合法合规小结=summary,
            编制日期="2026年2月26日"
        )

    def _parse_regulation(self) -> list:
        """解析法规政策Sheet"""
        logger.info("  解析法规政策...")
        sheet = self._get_sheet(self.SHEET_REGULATION)
        
        if sheet is None:
            return []
        
        from src.models.compliance_data import RegulationCompliance
        regulations = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            try:
                reg = RegulationCompliance(
                    法规名称=str(row[0]) if row[0] else "",
                    发布单位=str(row[1]) if len(row) > 1 and row[1] else None,
                    发布时间=str(row[2]) if len(row) > 2 and row[2] else None,
                    符合性分析=str(row[3]) if len(row) > 3 and row[3] else "",
                    符合性结论=str(row[4]) if len(row) > 4 and row[4] else ""
                )
                regulations.append(reg)
            except:
                continue
        
        return regulations

    def _parse_three_lines(self):
        """解析三线分析Sheet"""
        from src.models.compliance_data import ThreeLinesAnalysis
        sheet = self._get_sheet(self.SHEET_THREE_LINES)
        
        if sheet is None:
            return ThreeLinesAnalysis(
                是否占用耕地=False, 是否占用永久基本农田=False,
                是否占用生态保护红线=False, 是否占用城镇开发边界=False,
                符合性说明="Sheet不存在"
            )
        
        data = self._read_key_value_sheet(sheet)
        return ThreeLinesAnalysis(
            是否占用耕地=self._parse_bool(data.get("是否占用耕地", "否")),
            耕地面积=data.get("占用耕地面积（平方米）"),
            是否占用永久基本农田=self._parse_bool(data.get("是否占用永久基本农田", "否")),
            永久基本农田面积=data.get("占用永久基本农田面积（平方米）"),
            是否占用生态保护红线=self._parse_bool(data.get("是否占用生态保护红线", "否")),
            生态保护红线面积=data.get("占用生态保护红线面积（平方米）"),
            是否占用城镇开发边界=self._parse_bool(data.get("是否位于城镇开发边界内", "否")),
            城镇开发边界说明=data.get("城镇开发边界说明"),
            符合性说明=data.get("符合性说明", ""),
            数据来源=data.get("数据来源")
        )


    def _parse_spatial_planning(self):
        """解析国土空间规划Sheet"""
        from src.models.compliance_data import SpatialPlanningCompliance, OneMapAnalysis, FunctionalZoneAnalysis
        sheet = self._get_sheet(self.SHEET_SPATIAL_PLANNING)
        
        if sheet is None:
            return SpatialPlanningCompliance(
                一张图分析=OneMapAnalysis(是否上图=False, 落位说明=""),
                功能分区准入=FunctionalZoneAnalysis(城镇建设适宜性="", 生态保护重要性="", 农业生产适宜性="", 符合性说明=""),
                用途管制符合性="", 国土空间格局符合性="", 总体符合性结论=""
            )
        
        category_data = self._read_category_sheet(sheet)
        
        one_map_data = category_data.get("一张图落位", {})
        one_map = OneMapAnalysis(
            是否上图=one_map_data.get("是否上图落位", "是") == "是",
            重点项目库名称=one_map_data.get("重点项目库名称"),
            项目类型=one_map_data.get("项目类型"),
            落位说明=one_map_data.get("落位说明", "")
        )
        
        func_zone_data = category_data.get("功能分区准入", {})
        func_zone = FunctionalZoneAnalysis(
            城镇建设适宜性=func_zone_data.get("城镇建设适宜性", ""),
            生态保护重要性=func_zone_data.get("生态保护重要性", ""),
            农业生产适宜性=func_zone_data.get("农业生产适宜性", ""),
            符合性说明=func_zone_data.get("符合性说明", "")
        )
        
        usage_data = category_data.get("用途管制", {})
        spatial_data = category_data.get("总体格局", {})
        conclusion_data = category_data.get("总体结论", {})
        
        return SpatialPlanningCompliance(
            一张图分析=one_map,
            功能分区准入=func_zone,
            用途管制符合性=usage_data.get("符合性说明", ""),
            国土空间格局符合性=spatial_data.get("符合性说明", ""),
            总体符合性结论=conclusion_data.get("总体符合性结论", "")
        )


    def _parse_special_planning(self):
        """解析专项规划Sheet"""
        from src.models.compliance_data import SpecialPlanningCompliance, SpecialPlanCompliance
        sheet = self._get_sheet(self.SHEET_SPECIAL_PLANNING)
        
        if sheet is None:
            return SpecialPlanningCompliance(
                综合交通规划=SpecialPlanCompliance(规划名称="", 符合性分析="", 符合性结论=""),
                市政基础设施规划=SpecialPlanCompliance(规划名称="", 符合性分析="", 符合性结论=""),
                历史文化遗产保护规划=SpecialPlanCompliance(规划名称="", 符合性分析="", 符合性结论=""),
                综合防灾工程规划=SpecialPlanCompliance(规划名称="", 符合性分析="", 符合性结论=""),
                旅游规划=SpecialPlanCompliance(规划名称="", 符合性分析="", 符合性结论="")
            )
        
        special_plans = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            plan_type = str(row[0]).strip()
            special_plans[plan_type] = SpecialPlanCompliance(
                规划名称=str(row[1]) if len(row) > 1 and row[1] else "",
                符合性分析=str(row[2]) if len(row) > 2 and row[2] else "",
                符合性结论=str(row[3]) if len(row) > 3 and row[3] else ""
            )
        
        return SpecialPlanningCompliance(
            综合交通规划=special_plans.get("综合交通规划"),
            市政基础设施规划=special_plans.get("市政基础设施规划"),
            历史文化遗产保护规划=special_plans.get("历史文化遗产保护规划"),
            综合防灾工程规划=special_plans.get("综合防灾工程规划"),
            旅游规划=special_plans.get("旅游规划"),
            环境保护规划=special_plans.get("环境保护规划"),
            自然保护地规划=special_plans.get("自然保护地规划")
        )

    def _parse_other_planning(self):
        """解析其他规划Sheet"""
        from src.models.compliance_data import OtherPlanningCompliance, SpecialPlanCompliance
        sheet = self._get_sheet(self.SHEET_OTHER_PLANNING)
        
        if sheet is None:
            return OtherPlanningCompliance(
                国民经济和社会发展规划=SpecialPlanCompliance(规划名称="", 符合性分析="", 符合性结论=""),
                生态环境保护规划=SpecialPlanCompliance(规划名称="", 符合性分析="", 符合性结论=""),
                三线一单生态环境分区管控=SpecialPlanCompliance(规划名称="", 符合性分析="", 符合性结论="")
            )
        
        other_plans = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            plan_type = str(row[0]).strip()
            other_plans[plan_type] = SpecialPlanCompliance(
                规划名称=str(row[1]) if len(row) > 1 and row[1] else "",
                符合性分析=str(row[2]) if len(row) > 2 and row[2] else "",
                符合性结论=str(row[3]) if len(row) > 3 and row[3] else ""
            )
        
        return OtherPlanningCompliance(
            国民经济和社会发展规划=other_plans.get("国民经济和社会发展规划"),
            生态环境保护规划=other_plans.get("生态环境保护规划"),
            三线一单生态环境分区管控=other_plans.get("三线一单生态环境分区管控"),
            综合交通体系规划=other_plans.get("综合交通体系规划")
        )

    def _parse_urban_planning(self):
        """解析城乡总体规划Sheet"""
        from src.models.compliance_data import UrbanPlanningCompliance
        sheet = self._get_sheet(self.SHEET_URBAN_PLANNING)
        
        if sheet is None:
            return None
        
        data = self._read_key_value_sheet(sheet)
        return UrbanPlanningCompliance(
            规划名称=data.get("规划名称", ""),
            规划期限=data.get("规划期限", ""),
            空间管制分区=data.get("空间管制分区", ""),
            符合性分析=data.get("符合性分析", ""),
            符合性结论=data.get("符合性结论", "")
        )

    def _parse_compliance_summary(self) -> str:
        """解析合法合规小结"""
        sheet = self._get_sheet(self.SHEET_COMPLIANCE_SUMMARY)
        
        if sheet is None:
            return "Sheet 不存在"
        
        data = self._read_key_value_sheet(sheet)
        return data.get("合法合规小结", "")
    
    def parse_rationality(self) -> 'RationalityData':
        """
        解析选址合理性分析数据（第 4 章）
        
        Returns:
            RationalityData: 选址合理性分析数据模型
        """
        if RationalityData is None:
            raise ImportError("RationalityData 模型未导入，请检查 rationality_data 模块")
        
        logger.info("开始解析选址合理性分析数据...")
        
        # 获取项目基本信息
        project_basic = self.parse_project_overview().to_dict()
        
        # 解析各个部分
        environmental = self._parse_environmental_impact()
        mineral = self._parse_mineral_resources()
        geological = self._parse_geological_hazards()
        social = self._parse_social_stability()
        energy = self._parse_energy_saving()
        summary = self._parse_rationality_summary()
        
        logger.info("选址合理性分析数据解析完成")
        
        return RationalityData(
            项目基本信息=project_basic,
            环境影响分析=environmental,
            矿产资源压覆=mineral,
            地质灾害分析=geological,
            社会稳定分析=social,
            节能分析=energy,
            合理性结论=summary,
            数据来源="用户提供 Excel 数据",
            编制日期="2026 年 2 月 27 日"
        )
    
    def _parse_environmental_impact(self):
        """解析环境影响分析 Sheet"""
        if EnvironmentalImpactAnalysis is None:
            return None
        
        logger.info("  解析环境影响分析...")
        sheet = self._get_sheet(self.SHEET_ENVIRONMENTAL)
        
        if sheet is None:
            return EnvironmentalImpactAnalysis(
                大气环境影响=AtmosphericImpact(防治结论=" Sheet 不存在"),
                噪声环境影响=NoiseImpact(防治结论=" Sheet 不存在"),
                水环境影响=WaterImpact(防治结论=" Sheet 不存在"),
                固体废弃物影响=SolidWasteImpact(防治结论=" Sheet 不存在"),
                交通影响=TrafficImpact(防治结论=""),
                生态修复=EcologicalRestoration(修复措施=[], 防治结论=""),
                环境影响小结=""
            )
        
        category_data = self._read_category_sheet(sheet)
        
        # 大气环境影响
        air_data = category_data.get("大气环境", {})
        atmospheric = AtmosphericImpact(
            施工期扬尘措施=[v for k, v in air_data.items() if "施工期扬尘" in k or "扬尘" in k],
            施工机械废气措施=[v for k, v in air_data.items() if "废气" in k or "机械" in k],
            运营期废气措施=[v for k, v in air_data.items() if "运营期" in k and "废气" in k],
            影响程度=air_data.get("影响程度", "影响较小"),
            防治结论=air_data.get("防治结论", "")
        )
        
        # 噪声环境影响
        noise_data = category_data.get("噪声环境", {})
        noise = NoiseImpact(
            施工期噪声措施=[v for k, v in noise_data.items() if "施工期" in k or "噪声" in k],
            运营期噪声措施=[v for k, v in noise_data.items() if "运营期" in k and "噪声" in k],
            影响程度=noise_data.get("影响程度", "影响较小"),
            防治结论=noise_data.get("防治结论", "")
        )
        
        # 水环境影响
        water_data = category_data.get("水环境", {})
        water = WaterImpact(
            施工期废水措施=[v for k, v in water_data.items() if "施工期" in k or "废水" in k],
            运营期废水措施=[v for k, v in water_data.items() if "运营期" in k and "废水" in k],
            影响程度=water_data.get("影响程度", "影响较小"),
            防治结论=water_data.get("防治结论", "")
        )
        
        # 固体废弃物影响
        waste_data = category_data.get("固体废弃物", {})
        waste = SolidWasteImpact(
            施工期固废措施=[v for k, v in waste_data.items() if "施工期" in k or "固废" in k],
            运营期固废措施=[v for k, v in waste_data.items() if "运营期" in k and "固废" in k],
            影响程度=waste_data.get("影响程度", "影响较小"),
            防治结论=waste_data.get("防治结论", "")
        )
        
        # 交通影响
        traffic_data = category_data.get("交通影响", {})
        traffic = TrafficImpact(
            施工期交通影响=traffic_data.get("施工期交通影响", "施工期间交通影响较小，随工程结束而消失"),
            施工期缓解措施=[v for k, v in traffic_data.items() if "缓解" in k or "措施" in k],
            运营期交通影响=traffic_data.get("运营期交通影响"),
            运营期缓解措施=[],
            防治结论=traffic_data.get("防治结论", "交通影响较小，可接受")
        )
        
        # 生态修复
        eco_data = category_data.get("生态修复", {})
        ecological = EcologicalRestoration(
            对居民点影响=eco_data.get("对居民点影响", "影响较小"),
            居民点防治措施=[v for k, v in eco_data.items() if "居民点" in k and "措施" in k],
            对动物影响=eco_data.get("对动物影响", "影响较小"),
            动物防治措施=[v for k, v in eco_data.items() if "动物" in k and "措施" in k],
            对植物影响=eco_data.get("对植物影响", "影响较小"),
            植物防治措施=[v for k, v in eco_data.items() if "植物" in k and "措施" in k],
            水土保持措施=[v for k, v in eco_data.items() if "水土" in k],
            防治结论=eco_data.get("防治结论", "生态影响较小，可接受")
        )
        
        # 环境影响小结
        summary_data = category_data.get("环境影响小结", {})
        summary = summary_data.get("小结内容", "") if isinstance(summary_data, dict) else str(summary_data)
        
        return EnvironmentalImpactAnalysis(
            大气环境影响=atmospheric,
            噪声环境影响=noise,
            水环境影响=water,
            固体废弃物影响=waste,
            交通影响=traffic,
            生态修复=ecological,
            环境影响小结=summary
        )
    
    def _parse_mineral_resources(self):
        """解析矿产资源压覆 Sheet"""
        if MineralResourceAnalysis is None:
            return None
        
        logger.info("  解析矿产资源压覆...")
        sheet = self._get_sheet(self.SHEET_MINERAL)
        
        if sheet is None:
            return MineralResourceAnalysis(
                是否压覆矿产资源=False,
                是否与采矿权重叠=False,
                是否与探矿权重叠=False,
                是否与地质项目重叠=False,
                分析结论="Sheet 不存在"
            )
        
        data = self._read_key_value_sheet(sheet)
        return MineralResourceAnalysis(
            是否压覆矿产资源=self._parse_bool(data.get("是否压覆矿产资源", "否")),
            是否与采矿权重叠=self._parse_bool(data.get("是否与采矿权重叠", "否")),
            是否与探矿权重叠=self._parse_bool(data.get("是否与探矿权重叠", "否")),
            是否与地质项目重叠=self._parse_bool(data.get("是否与地质项目重叠", "否")),
            复函信息=data.get("复函信息"),
            分析结论=data.get("分析结论", "")
        )
    
    def _parse_geological_hazards(self):
        """解析地质灾害影响 Sheet"""
        if GeologicalHazardAnalysis is None:
            return None
        
        logger.info("  解析地质灾害影响...")
        sheet = self._get_sheet(self.SHEET_GEOLOGICAL)
        
        if sheet is None:
            return GeologicalHazardAnalysis(
                地质灾害类型=[],
                地质灾害易发程度="低易发区",
                危险性等级="小",
                地震基本烈度="6 度",
                防治措施=[],
                分析结论="Sheet 不存在"
            )
        
        data = self._read_key_value_sheet(sheet)
        
        # 解析灾害类型（逗号分隔）
        hazard_types_str = data.get("地质灾害类型", "")
        if isinstance(hazard_types_str, str) and hazard_types_str:
            hazard_types = [t.strip() for t in hazard_types_str.split(",") if t.strip()]
        else:
            hazard_types = []
        
        # 解析防治措施（逗号分隔）
        prevention_str = data.get("防治措施", "")
        if isinstance(prevention_str, str) and prevention_str:
            prevention_measures = [m.strip() for m in prevention_str.split(",") if m.strip()]
        else:
            prevention_measures = []
        
        return GeologicalHazardAnalysis(
            地质灾害类型=hazard_types,
            地质灾害易发程度=data.get("地质灾害易发程度", "低易发区"),
            危险性等级=data.get("危险性等级", "小"),
            地震基本烈度=data.get("地震基本烈度", "6 度"),
            地震动峰值加速度=data.get("地震动峰值加速度"),
            防治措施=prevention_measures,
            分析结论=data.get("分析结论", "")
        )
    
    def _parse_social_stability(self):
        """解析社会稳定影响 Sheet"""
        if SocialStabilityAnalysis is None:
            return None
        
        logger.info("  解析社会稳定影响...")
        sheet = self._get_sheet(self.SHEET_SOCIAL_STABILITY)
        
        if sheet is None:
            return SocialStabilityAnalysis(
                合法性风险=LegalityRiskAnalysis(风险内容="", 风险等级="低", 防范措施=[]),
                生活环境风险=LivingEnvironmentRisk(风险内容="", 风险等级="低", 防范措施=[]),
                社会环境风险=SocialEnvironmentRisk(风险内容="", 风险等级="低", 防范措施=[]),
                社会稳定小结=""
            )
        
        category_data = self._read_category_sheet(sheet)
        
        # 合法性风险
        legality_data = category_data.get("合法性风险", {})
        legality = LegalityRiskAnalysis(
            风险内容=legality_data.get("风险内容", ""),
            风险等级=legality_data.get("风险等级", "低"),
            防范措施=[v for k, v in legality_data.items() if "防范" in k or "措施" in k]
        )
        
        # 生活环境风险
        living_data = category_data.get("生活环境风险", {})
        living = LivingEnvironmentRisk(
            风险内容=living_data.get("风险内容", ""),
            风险等级=living_data.get("风险等级", "低"),
            防范措施=[v for k, v in living_data.items() if "防范" in k or "措施" in k]
        )
        
        # 社会环境风险
        social_data = category_data.get("社会环境风险", {})
        social = SocialEnvironmentRisk(
            风险内容=social_data.get("风险内容", ""),
            风险等级=social_data.get("风险等级", "低"),
            防范措施=[v for k, v in social_data.items() if "防范" in k or "措施" in k]
        )
        
        # 社会稳定小结
        summary_data = category_data.get("社会稳定小结", {})
        summary = summary_data.get("小结内容", "") if isinstance(summary_data, dict) else str(summary_data) if summary_data else ""
        
        return SocialStabilityAnalysis(
            合法性风险=legality,
            生活环境风险=living,
            社会环境风险=social,
            社会稳定小结=summary
        )
    
    def _parse_energy_saving(self):
        """解析节能分析 Sheet"""
        if EnergySavingAnalysis is None:
            return None
        
        logger.info("  解析节能分析...")
        sheet = self._get_sheet(self.SHEET_ENERGY_SAVING)
        
        if sheet is None:
            return EnergySavingAnalysis(
                节能标准=[],
                节能措施=[],
                能耗指标="",
                节能结论=""
            )
        
        data = self._read_key_value_sheet(sheet)
        
        # 解析节能标准（逗号分隔）
        standards_str = data.get("节能标准", "")
        if isinstance(standards_str, str) and standards_str:
            standards = [s.strip() for s in standards_str.split(",") if s.strip()]
        else:
            standards = []
        
        # 解析节能措施（逗号分隔）
        measures_str = data.get("节能措施", "")
        if isinstance(measures_str, str) and measures_str:
            measures = [m.strip() for m in measures_str.split(",") if m.strip()]
        else:
            measures = []
        
        return EnergySavingAnalysis(
            节能标准=standards,
            节能措施=measures,
            能耗指标=data.get("能耗指标", ""),
            节能结论=data.get("节能结论", "")
        )
    
    def _parse_rationality_summary(self) -> str:
        """解析合理性小结"""
        sheet = self._get_sheet(self.SHEET_RATIONALITY_SUMMARY)
        
        if sheet is None:
            return "Sheet 不存在"
        
        data = self._read_key_value_sheet(sheet)
        return data.get("合理性结论", "")

    def parse_land_use(self) -> 'LandUseData':
        """
        解析节约集约用地分析数据（第 5 章）
        
        Returns:
            LandUseData: 节约集约用地分析数据模型
        """
        if LandUseData is None:
            raise ImportError("LandUseData 模型未导入，请检查 land_use_data 模块")
        
        logger.info("开始解析节约集约用地分析数据...")
        
        # 获取项目基本信息
        project_basic = self.parse_project_overview().to_dict()
        
        # 解析各个部分
        functional_zones = self._parse_functional_zones()
        land_scale = self._parse_land_scale()
        land_tech = self._parse_land_tech()
        case_comparison = self._parse_case_comparison()
        summary = self._parse_land_use_summary()
        
        logger.info("节约集约用地分析数据解析完成")
        
        return LandUseData(
            项目基本信息=project_basic,
            功能分区情况=functional_zones,
            用地规模合理性=land_scale,
            采用的节地技术=land_tech,
            案例对比情况=case_comparison,
            节约集约用地小结=summary,
            数据来源="用户提供 Excel 数据",
            编制日期="2026 年 2 月 27 日"
        )
    
    def _parse_functional_zones(self) -> list:
        """解析功能分区 Sheet"""
        if FunctionalZone is None:
            return []
        
        logger.info("  解析功能分区...")
        sheet = self._get_sheet(self.SHEET_FUNCTIONAL_ZONE)
        
        if sheet is None:
            return [
                FunctionalZone(
                    分区名称="生产区用地",
                    分区面积=0.0,
                    占比=0.0,
                    功能描述="Sheet 不存在"
                )
            ]
        
        zones = []
        rows = self._read_table_sheet(sheet)
        
        for row in rows:
            try:
                # 解析子分区（如果有）
                子分区 = None
                if row.get("子分区"):
                    子分区_str = str(row.get("子分区", ""))
                    子分区 = [{"分区名称": s.strip()} for s in 子分区_str.split(",") if s.strip()]
                
                zone = FunctionalZone(
                    分区名称=str(row.get("分区名称", "")),
                    分区面积=float(row.get("分区面积", 0) or 0),
                    占比=float(row.get("占比", 0) or 0),
                    子分区=子分区,
                    功能描述=row.get("功能描述"),
                    用地依据=row.get("用地依据")
                )
                zones.append(zone)
            except Exception as e:
                logger.warning(f"解析功能分区行失败: {str(e)}")
                continue
        
        if not zones:
            zones = [FunctionalZone(分区名称="未分类", 分区面积=1.0, 占比=100.0)]
        logger.info(f"  解析到 {len(zones)} 个功能分区")
        return zones
    
    def _parse_land_scale(self):
        """解析用地规模合理性 Sheet"""
        if LandScaleAnalysis is None:
            return None
        
        logger.info("  解析用地规模合理性...")
        sheet = self._get_sheet(self.SHEET_LAND_SCALE)
        
        if sheet is None:
            return LandScaleAnalysis(
                总体指标=OverallLandIndicator(
                    项目总用地面积=1.0,
                    建设规模="",
                    标准依据="",
                    标准要求范围="",
                    是否符合要求=True
                ),
                各分区指标=[]
            )
        
        category_data = self._read_category_sheet(sheet)
        
        # 解析总体指标
        overall_data = category_data.get("总体指标", {})
        overall = OverallLandIndicator(
            项目总用地面积=float(overall_data.get("项目总用地面积", 1) or 1),
            建设规模=str(overall_data.get("建设规模", "")),
            标准依据=str(overall_data.get("标准依据", "")),
            标准要求范围=str(overall_data.get("标准要求范围", "")),
            是否符合要求=self._parse_bool(overall_data.get("是否符合要求", "是")),
            对比分析=overall_data.get("对比分析")
        )
        
        # 解析各分区指标
        sub_indicators = []
        sub_areas_data = category_data.get("各分区指标", {})
        for area_name, area_data in sub_areas_data.items():
            if isinstance(area_data, dict):
                sub_indicators.append(SubAreaIndicator(
                    区域名称=area_name,
                    实际用地面积=float(area_data.get("实际用地面积", 0) or 0),
                    标准依据=str(area_data.get("标准依据", "")),
                    标准指标值=str(area_data.get("标准指标值", "")),
                    是否符合要求=self._parse_bool(area_data.get("是否符合要求", "是")),
                    对比分析=area_data.get("对比分析")
                ))
        
        # 辅助区用地占比
        aux_data = category_data.get("辅助区用地占比", {})
        辅助区用地占比 = {
            "实际占比": aux_data.get("实际占比", ""),
            "标准要求范围": aux_data.get("标准要求范围", ""),
            "是否符合": self._parse_bool(aux_data.get("是否符合", "是"))
        } if aux_data else None
        
        # 综合评价
        summary_data = category_data.get("综合评价", {})
        综合评价 = summary_data.get("综合评价", "") if isinstance(summary_data, dict) else str(summary_data) if summary_data else None
        
        return LandScaleAnalysis(
            总体指标=overall,
            各分区指标=sub_indicators,
            辅助区用地占比=辅助区用地占比,
            综合评价=综合评价
        )
    
    def _parse_land_tech(self):
        """解析节地技术 Sheet"""
        if LandSavingTechnology is None:
            return None
        
        logger.info("  解析节地技术...")
        sheet = self._get_sheet(self.SHEET_LAND_TECH)
        
        if sheet is None:
            return LandSavingTechnology(
                前期工作阶段措施=[],
                建设实施阶段措施=[]
            )
        
        category_data = self._read_category_sheet(sheet)
        
        # 解析前期工作阶段措施
        pre_construction = []
        pre_data = category_data.get("前期工作阶段", {})
        for name, desc in pre_data.items():
            if isinstance(desc, str):
                pre_construction.append(PreConstructionMeasure(
                    措施名称=name,
                    措施描述=desc
                ))
            elif isinstance(desc, dict):
                pre_construction.append(PreConstructionMeasure(
                    措施名称=name,
                    措施描述=desc.get("措施描述", ""),
                    实施效果=desc.get("实施效果")
                ))
        
        # 解析建设实施阶段措施
        construction_phase = []
        const_data = category_data.get("建设实施阶段", {})
        for name, desc in const_data.items():
            if isinstance(desc, str):
                construction_phase.append(ConstructionPhaseMeasure(
                    措施名称=name,
                    措施描述=desc
                ))
            elif isinstance(desc, dict):
                construction_phase.append(ConstructionPhaseMeasure(
                    措施名称=name,
                    措施描述=desc.get("措施描述", ""),
                    实施主体=desc.get("实施主体"),
                    实施效果=desc.get("实施效果")
                ))
        
        # 综合评价
        summary_data = category_data.get("综合评价", {})
        综合评价 = summary_data.get("综合评价", "") if isinstance(summary_data, dict) else str(summary_data) if summary_data else None
        
        return LandSavingTechnology(
            前期工作阶段措施=pre_construction if pre_construction else [],
            建设实施阶段措施=construction_phase if construction_phase else [],
            综合评价=综合评价
        )
    
    def _parse_case_comparison(self):
        """解析案例对比 Sheet"""
        if CaseComparison is None:
            return None
        
        logger.info("  解析案例对比...")
        sheet = self._get_sheet(self.SHEET_CASE_COMPARISON)
        
        if sheet is None:
            return CaseComparison(
                本项目=ComparisonCase(
                    案例名称="本项目",
                    建设规模="",
                    用地面积=1.0,
                    总投资=1.0
                ),
                对比案例=[],
                对比结论="Sheet 不存在"
            )
        
        category_data = self._read_category_sheet(sheet)
        
        # 解析本项目
        project_data = category_data.get("本项目", {})
        本项目 = ComparisonCase(
            案例名称="本项目",
            建设规模=str(project_data.get("建设规模", "")),
            用地面积=float(project_data.get("用地面积", 1) or 1),
            总投资=float(project_data.get("总投资", 1) or 1),
            采用技术=project_data.get("采用技术")
        )
        
        # 解析对比案例
        comparison_cases = []
        for key, value in category_data.items():
            if key.startswith("案例") and isinstance(value, dict):
                comparison_cases.append(ComparisonCase(
                    案例名称=value.get("案例名称", key),
                    案例地点=value.get("案例地点"),
                    建设规模=str(value.get("建设规模", "")),
                    用地面积=float(value.get("用地面积", 1) or 1),
                    总投资=float(value.get("总投资", 1) or 1),
                    采用技术=value.get("采用技术"),
                    数据来源=value.get("数据来源")
                ))
        
        # 解析对比案例
        comparison_cases = []
        for key, value in category_data.items():
            if key.startswith("案例") and isinstance(value, dict):
                comparison_cases.append(ComparisonCase(
                    案例名称=value.get("案例名称", key),
                    案例地点=value.get("案例地点"),
                    建设规模=str(value.get("建设规模", "")),
                    用地面积=float(value.get("用地面积", 0) or 0),
                    总投资=float(value.get("总投资", 0) or 0),
                    采用技术=value.get("采用技术"),
                    数据来源=value.get("数据来源")
                ))
        
        # 解析对比结论
        conclusion_data = category_data.get("对比结论", {})
        对比结论 = conclusion_data.get("对比结论", "") if isinstance(conclusion_data, dict) else str(conclusion_data) if conclusion_data else ""
        
        # 解析单位投资对比
        unit_data = category_data.get("单位投资对比", {})
        单位投资对比 = unit_data if isinstance(unit_data, dict) else None
        
        # 如果没有对比案例，添加一个默认案例
        if not comparison_cases:
            comparison_cases = [ComparisonCase(
                案例名称="参考案例1",
                案例地点="待补充",
                建设规模="待补充",
                用地面积=1.0,
                总投资=1.0
            )]
        
        return CaseComparison(
            本项目=本项目,
            对比案例=comparison_cases,
            对比结论=对比结论,
            单位投资对比=单位投资对比
        )
    
    def _parse_land_use_summary(self) -> str:
        """解析节约集约用地小结"""
        sheet = self._get_sheet(self.SHEET_LAND_USE_SUMMARY)
        
        if sheet is None:
            return "Sheet 不存在"
        
        data = self._read_key_value_sheet(sheet)
        return data.get("节约集约用地小结", "")
    
    def parse_conclusion(self) -> 'ConclusionData':
        """
        解析结论与建议数据（第 6 章）
        
        Returns:
            ConclusionData: 结论与建议数据模型
        """
        if ConclusionData is None:
            raise ImportError("ConclusionData 模型未导入，请检查 conclusion_data 模块")
        
        logger.info("开始解析结论与建议数据...")
        
        # 获取项目基本信息
        project_basic = self.parse_project_overview().to_dict()
        
        # 解析结论建议Sheet
        sheet = self._get_sheet(self.SHEET_CONCLUSION)
        
        if sheet is None:
            # 如果Sheet不存在，返回示例数据
            logger.warning("结论建议 Sheet 不存在，使用默认数据")
            from src.models.conclusion_data import get_sample_data
            return get_sample_data()
        
        data = self._read_key_value_sheet(sheet)
        
        # 构建合规性结论
        合法合规性结论 = 合规性结论(
            法律法规结论=data.get("法律法规结论", "符合相关法律法规及政策文件"),
            三线结论={
                "耕地和永久基本农田": data.get("耕地和永久基本农田结论", "不占用"),
                "生态保护红线": data.get("生态保护红线结论", "不占用"),
                "城镇开发边界": data.get("城镇开发边界结论", "不占用"),
            },
            国土空间规划结论={
                "一张图上图落位情况": data.get("一张图上图落位情况", "已上图"),
                "功能分区准入": data.get("功能分区准入", "符合"),
                "用途管制": data.get("用途管制", "符合"),
            },
            专项规划结论={
                "综合交通规划": data.get("综合交通规划符合性", "符合"),
                "市政基础设施规划": data.get("市政基础设施规划符合性", "符合"),
                "历史文化遗产保护规划": data.get("历史文化遗产保护规划符合性", "符合"),
            },
            其他规划结论={
                "国民经济和社会发展规划": data.get("国民经济规划符合性", "符合"),
            } if data.get("国民经济规划符合性") else None,
            城乡总体规划结论=data.get("城乡总体规划结论"),
            综合结论=data.get("合法合规综合结论", "符合相关法律法规及各项规划要求"),
        )
        
        # 构建合理性结论
        选址合理性结论 = 合理性结论(
            环境影响结论=data.get("环境影响结论", "影响较小，有防治措施"),
            矿产资源结论=data.get("矿产资源结论", "未压覆已查明矿产"),
            地质灾害结论=data.get("地质灾害结论", "受灾程度小，有防治措施"),
            社会稳定结论=data.get("社会稳定结论"),
            节能结论=data.get("节能结论"),
            综合结论=data.get("选址合理性综合结论", "选址合理，符合项目建设要求"),
        )
        
        # 构建节约集约用地结论
        节约集约用地结论 = 节约集约结论(
            功能分区结论=data.get("功能分区结论", "功能分区合理"),
            用地规模结论=data.get("用地规模结论", "用地规模各分区合理"),
            节地技术结论=data.get("节地技术结论", "较其他项目节地水平更高"),
            综合结论=data.get("节约集约用地综合结论", "符合节约集约用地要求"),
        )
        
        # 构建建议列表（固定5条）
        建议列表 = []
        for i in range(1, 6):
            建议 = data.get(f"建议{i}", "")
            if 建议:
                建议列表.append(建议项(序号=i, 内容=建议))
        
        # 如果Excel中没有足够的建议，使用默认建议
        if len(建议列表) < 5:
            默认建议 = [
                "项目应进一步衔接和协调建设项目选址与各类规划的关系。",
                "项目选址应进一步分析项目运营对周边敏感目标的潜在环境影响，确保满足卫生防护距离要求。",
                "项目选址须精准落入城镇排水与污水处理系统专项规划所确定的设施布局和服务范围内。",
                "项目选址应精确核算厂区防洪排涝标准，确保厂区不受洪涝威胁。",
                "项目在环评基础上，可针对选址特性，深化恶臭气体扩散模拟，精确划定卫生防护距离。",
            ]
            for i, 默认 in enumerate(默认建议, 1):
                if i > len(建议列表):
                    建议列表.append(建议项(序号=i, 内容=默认))
        
        logger.info("结论与建议数据解析完成")
        
        return ConclusionData(
            项目基本信息=project_basic,
            合法合规性结论=合法合规性结论,
            选址合理性结论=选址合理性结论,
            节约集约用地结论=节约集约用地结论,
            综合论证结论=data.get("综合论证结论", "该项目选址符合项目所需的建设场地要求，与各项规划、政策基本协调，项目选址可行"),
            建议列表=建议列表,
            数据来源="用户提供 Excel 数据",
            编制日期="2026 年 2 月 27 日"
        )
    
    def parse_all(self) -> Tuple[ProjectOverviewData, SiteSelectionData]:
        """
        解析所有数据
        
        Returns:
            (ProjectOverviewData, SiteSelectionData) 元组
        """
        logger.info(f"开始解析Excel文件: {self.file_path}")
        
        project_overview = self.parse_project_overview()
        site_selection = self.parse_site_selection()
        logger.info("Excel文件解析完成")
        return project_overview, site_selection
    def parse_all_with_chapter3(self):
        """解析所有数据（包括第 3 章）"""
        logger.info(f"开始解析 Excel 文件：{self.file_path}")
        project_overview = self.parse_project_overview()
        site_selection = self.parse_site_selection()
        compliance = self.parse_compliance()
        logger.info("Excel 文件解析完成")
        return project_overview, site_selection, compliance
    
    def parse_all_with_chapter6(self):
        """解析所有数据（包括第 3、4、5、6 章）"""
        logger.info(f"开始解析 Excel 文件：{self.file_path}")
        project_overview = self.parse_project_overview()
        site_selection = self.parse_site_selection()
        compliance = self.parse_compliance()
        rationality = self.parse_rationality()
        land_use = self.parse_land_use()
        conclusion = self.parse_conclusion()
        logger.info("Excel 文件解析完成")
        return project_overview, site_selection, compliance, rationality, land_use, conclusion
    
    def validate_data(self, sheets: List[str] = None) -> 'ValidationReport':
        """
        验证Excel数据完整性
        
        使用 DataValidator 验证数据，返回验证报告。
        
        Args:
            sheets: 要验证的Sheet列表，None则验证所有
            
        Returns:
            ValidationReport: 验证报告
        
        Example:
            >>> parser = ExcelParser("项目数据.xlsx")
            >>> report = parser.validate_data()
            >>> print(report.to_markdown())
        """
        validator = DataValidator()
        return validator.validate_all(self, sheets)
    
    def fill_missing_data(
        self,
        output_path: str = None,
        fill_value: str = "待补充"
    ) -> Dict[str, int]:
        """
        填充缺失字段
        
        使用 DataValidator 填充缺失字段。
        
        Args:
            output_path: 输出路径，None则覆盖原文件
            fill_value: 填充值
            
        Returns:
            填充字段统计 {sheet_name: filled_count}
        
        Example:
            >>> parser = ExcelParser("项目数据.xlsx")
            >>> stats = parser.fill_missing_data("填充后.xlsx")
            >>> print(f"填充了 {sum(stats.values())} 个字段")
        """
        validator = DataValidator()
        return validator.fill_missing_fields(self, output_path, fill_value)
    
    def get_missing_fields(self) -> Dict[str, List[str]]:
        """
        获取缺失字段列表
        
        快捷方法，返回按Sheet分组的缺失字段。
        
        Returns:
            {sheet_name: [missing_fields]}
        
        Example:
            >>> parser = ExcelParser("项目数据.xlsx")
            >>> missing = parser.get_missing_fields()
            >>> for sheet, fields in missing.items():
            ...     print(f"{sheet}: {fields}")
        """
        report = self.validate_data()
        return report.get_missing_fields()
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

def parse_excel_data(file_path: str) -> Tuple[ProjectOverviewData, SiteSelectionData]:
    """
    便捷函数：解析Excel数据文件

    Args:
        file_path: Excel文件路径

    Returns:
        (ProjectOverviewData, SiteSelectionData) 元组
    """
    parser = ExcelParser(file_path)
    try:
        return parser.parse_all()
    finally:
        parser.close()


# 测试代码
if __name__ == "__main__":
    import sys

    # 查找模板文件
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    template_path = os.path.join(project_root, "templates", "excel_templates", "项目数据模板.xlsx")

    if len(sys.argv) > 1:
        template_path = sys.argv[1]

    print(f"测试Excel解析器...")
    print(f"模板路径: {template_path}")

    if not os.path.exists(template_path):
        print(f"模板文件不存在，请先创建模板文件")
        print(f"预期路径: {template_path}")
        sys.exit(0)

    try:
        parser = ExcelParser(template_path)

        # 测试解析项目基本信息
        print("\n=== 测试解析项目基本信息 ===")
        project_data = parser.parse_project_overview()
        print(f"项目名称: {project_data.项目名称}")
        print(f"建设单位: {project_data.建设单位}")
        print(f"项目投资: {project_data.项目投资}")

        # 测试解析选址数据
        print("\n=== 测试解析选址数据 ===")
        site_data = parser.parse_site_selection()
        print(f"备选方案数: {len(site_data.备选方案)}")
        print(f"征求意见数: {len(site_data.征求意见情况)}")
        print(f"推荐方案: {site_data.方案比选.推荐方案}")

        parser.close()
        print("\n测试通过!")

    except Exception as e:
        print(f"测试失败: {str(e)}")
        import traceback
        traceback.print_exc()