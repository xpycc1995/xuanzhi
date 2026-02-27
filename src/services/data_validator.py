"""
数据验证模块 - 基于Pydantic的Excel数据验证

正确的架构：
1. Excel → ExcelParser → 字典数据
2. 字典数据 → DataValidator → Pydantic验证
3. 验证结果 → ValidationReport (缺失/错误字段)
4. 完整数据 → Agent生成报告

不再让LLM"发现"空白字段，而是用模型进行严格验证。
"""

import json
import logging
from typing import Dict, List, Any, Optional, Tuple, Type
from dataclasses import dataclass, field
from datetime import datetime
from pydantic import BaseModel, ValidationError

logger = logging.getLogger(__name__)


@dataclass
class FieldValidationResult:
    """字段验证结果"""
    field_name: str
    status: str  # "valid", "missing", "invalid", "empty"
    value: Any = None
    error_message: Optional[str] = None
    required: bool = True
    

@dataclass
class SheetValidationResult:
    """Sheet验证结果"""
    sheet_name: str
    total_fields: int = 0
    valid_fields: int = 0
    missing_fields: int = 0
    invalid_fields: int = 0
    empty_fields: int = 0
    field_results: List[FieldValidationResult] = field(default_factory=list)
    
    @property
    def is_complete(self) -> bool:
        """数据是否完整"""
        return self.missing_fields == 0 and self.invalid_fields == 0
    
    @property
    def completion_rate(self) -> float:
        """完整率"""
        if self.total_fields == 0:
            return 0.0
        return self.valid_fields / self.total_fields * 100


@dataclass
class ValidationReport:
    """完整验证报告"""
    file_name: str
    validation_time: str
    total_sheets: int = 0
    total_fields: int = 0
    valid_fields: int = 0
    missing_fields: int = 0
    invalid_fields: int = 0
    sheet_results: List[SheetValidationResult] = field(default_factory=list)
    
    @property
    def is_complete(self) -> bool:
        """所有数据是否完整"""
        return all(sr.is_complete for sr in self.sheet_results)
    
    @property
    def completion_rate(self) -> float:
        """整体完整率"""
        if self.total_fields == 0:
            return 0.0
        return self.valid_fields / self.total_fields * 100
    
    def get_missing_fields(self) -> Dict[str, List[str]]:
        """获取所有缺失字段，按Sheet分组"""
        result = {}
        for sr in self.sheet_results:
            missing = [fr.field_name for fr in sr.field_results 
                      if fr.status in ("missing", "empty")]
            if missing:
                result[sr.sheet_name] = missing
        return result
    
    def to_json(self) -> str:
        """转换为JSON"""
        return json.dumps({
            "file_name": self.file_name,
            "validation_time": self.validation_time,
            "summary": {
                "total_sheets": self.total_sheets,
                "total_fields": self.total_fields,
                "valid_fields": self.valid_fields,
                "missing_fields": self.missing_fields,
                "completion_rate": f"{self.completion_rate:.1f}%",
                "is_complete": self.is_complete,
            },
            "missing_fields_by_sheet": self.get_missing_fields(),
            "sheets": [
                {
                    "name": sr.sheet_name,
                    "total": sr.total_fields,
                    "valid": sr.valid_fields,
                    "missing": sr.missing_fields,
                    "completion_rate": f"{sr.completion_rate:.1f}%",
                    "is_complete": sr.is_complete,
                }
                for sr in self.sheet_results
            ]
        }, ensure_ascii=False, indent=2)
    
    def to_markdown(self) -> str:
        """转换为Markdown报告"""
        lines = [
            f"# 数据验证报告",
            f"",
            f"**文件**: {self.file_name}",
            f"**时间**: {self.validation_time}",
            f"",
            f"## 总体情况",
            f"",
            f"| 指标 | 数值 |",
            f"|------|------|",
            f"| 总Sheet数 | {self.total_sheets} |",
            f"| 总字段数 | {self.total_fields} |",
            f"| 有效字段 | {self.valid_fields} |",
            f"| 缺失字段 | {self.missing_fields} |",
            f"| 完整率 | {self.completion_rate:.1f}% |",
            f"",
        ]
        
        # Sheet详情
        lines.append("## 各Sheet验证结果")
        lines.append("")
        lines.append("| Sheet名称 | 总字段 | 有效 | 缺失 | 完整率 | 状态 |")
        lines.append("|-----------|--------|------|------|--------|------|")
        
        for sr in self.sheet_results:
            status = "✅" if sr.is_complete else "⚠️"
            lines.append(
                f"| {sr.sheet_name} | {sr.total_fields} | {sr.valid_fields} | "
                f"{sr.missing_fields} | {sr.completion_rate:.1f}% | {status} |"
            )
        
        # 缺失字段详情
        missing = self.get_missing_fields()
        if missing:
            lines.append("")
            lines.append("## 缺失字段详情")
            lines.append("")
            
            for sheet_name, fields in missing.items():
                lines.append(f"### {sheet_name}")
                for f in fields:
                    lines.append(f"- {f}")
                lines.append("")
        
        return "\n".join(lines)


class DataValidator:
    """
    数据验证器
    
    使用Pydantic模型验证Excel数据，生成验证报告。
    
    使用方式:
    ```python
    from src.services.excel_parser import ExcelParser
    from src.services.data_validator import DataValidator
    
    # 读取Excel
    parser = ExcelParser("项目数据.xlsx")
    
    # 验证数据
    validator = DataValidator()
    
    # 验证单个Sheet
    result = validator.validate_sheet(parser, "项目基本信息")
    
    # 验证所有章节
    report = validator.validate_all(parser)
    
    # 输出报告
    print(report.to_markdown())
    ```
    """
    
    # 必填字段定义（按Sheet）- 涵盖全部6章
    REQUIRED_FIELDS = {
        # 第1章：项目概况
        "项目基本信息": [
            "项目名称", "建设单位", "建设性质", "项目投资", 
            "项目选址", "建设内容"
        ],
        
        # 第2章：选址分析
        "备选方案": [
            "方案编号", "方案名称", "位置", "面积"
        ],
        "场地条件": [
            "地形地貌", "气候", "水文地质条件"
        ],
        "敏感条件": [
            "是否涉及生态保护红线", "是否占用耕地"
        ],
        "施工运营": [
            "方案一总投资", "方案二总投资"
        ],
        "征求意见": [
            "部门", "日期", "结论"
        ],
        "方案比选": [
            "推荐方案", "推荐理由"
        ],
        
        # 第3章：合法合规性分析
        "三线分析": [
            "是否占用耕地", "是否占用生态保护红线", "符合性说明"
        ],
        "国土空间规划": [
            "是否上图落位", "落位说明"
        ],
        "法规政策": [
            "法规名称", "符合性分析"
        ],
        "专项规划": [
            "规划类型", "符合性结论"
        ],
        
        # 第4章：选址合理性分析
        "环境影响": [
            "影响程度", "防治结论"
        ],
        "地质灾害": [
            "地质灾害易发程度", "危险性等级"
        ],
        "社会稳定": [
            "风险等级", "防范措施"
        ],
        "节能分析": [
            "节能标准", "节能措施"
        ],
        
        # 第5章：节约集约用地
        "功能分区": [
            "分区名称", "分区面积"
        ],
        "用地规模": [
            "总用地面积", "建筑系数"
        ],
        "节地技术": [
            "技术名称", "应用效果"
        ],
        
        # 第6章：结论与建议
        "结论建议": [
            "总体结论"
        ],
    }
    
    # 字段默认值（用于缺失字段）
    DEFAULT_VALUES = {
        "建设单位": "待补充",
        "选址原则": "待补充",
        "符合性说明": "待补充",
        "规划名称": "待补充",
        "复函信息": "待补充",
        "地质灾害类型": "待补充",
        "建议1": "待补充",
        "建议2": "待补充",
        "建议3": "待补充",
        "建议4": "待补充",
        "建议5": "待补充",
    }
    
    def __init__(self):
        """初始化验证器"""
        pass
    
    def validate_sheet(
        self,
        parser,  # ExcelParser实例
        sheet_name: str,
        required_fields: Optional[List[str]] = None,
    ) -> SheetValidationResult:
        """
        验证单个Sheet的数据
        
        Args:
            parser: ExcelParser实例
            sheet_name: Sheet名称
            required_fields: 必填字段列表，默认使用REQUIRED_FIELDS
            
        Returns:
            SheetValidationResult
        """
        sheet = parser._get_sheet(sheet_name)
        
        if sheet is None:
            return SheetValidationResult(
                sheet_name=sheet_name,
                total_fields=0,
                valid_fields=0,
                missing_fields=0,
                invalid_fields=0,
                field_results=[
                    FieldValidationResult(
                        field_name="Sheet",
                        status="missing",
                        error_message=f"Sheet '{sheet_name}' 不存在"
                    )
                ]
            )
        
        # 读取数据
        data = parser._read_key_value_sheet(sheet)
        
        # 获取必填字段
        if required_fields is None:
            required_fields = self.REQUIRED_FIELDS.get(sheet_name, [])
        
        # 验证每个字段
        results = []
        for field_name in required_fields:
            value = data.get(field_name)
            
            if value is None or str(value).strip() == "":
                # 字段缺失或为空
                results.append(FieldValidationResult(
                    field_name=field_name,
                    status="empty" if field_name in data else "missing",
                    value=None,
                    error_message=f"字段 '{field_name}' 为空或不存在",
                    required=True
                ))
            else:
                # 字段有效
                results.append(FieldValidationResult(
                    field_name=field_name,
                    status="valid",
                    value=value,
                    required=True
                ))
        
        # 统计
        total = len(results)
        valid = sum(1 for r in results if r.status == "valid")
        missing = sum(1 for r in results if r.status in ("missing", "empty"))
        
        return SheetValidationResult(
            sheet_name=sheet_name,
            total_fields=total,
            valid_fields=valid,
            missing_fields=missing,
            invalid_fields=0,
            field_results=results
        )
    
    def validate_all(
        self,
        parser,  # ExcelParser实例
        sheets: Optional[List[str]] = None,
    ) -> ValidationReport:
        """
        验证所有Sheet
        
        Args:
            parser: ExcelParser实例
            sheets: 要验证的Sheet列表，默认验证所有已定义的
            
        Returns:
            ValidationReport
        """
        if sheets is None:
            sheets = list(self.REQUIRED_FIELDS.keys())
        
        sheet_results = []
        for sheet_name in sheets:
            result = self.validate_sheet(parser, sheet_name)
            sheet_results.append(result)
        
        # 汇总统计
        total_fields = sum(sr.total_fields for sr in sheet_results)
        valid_fields = sum(sr.valid_fields for sr in sheet_results)
        missing_fields = sum(sr.missing_fields for sr in sheet_results)
        
        return ValidationReport(
            file_name=parser.file_path,
            validation_time=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            total_sheets=len(sheet_results),
            total_fields=total_fields,
            valid_fields=valid_fields,
            missing_fields=missing_fields,
            sheet_results=sheet_results
        )
    
    def validate_pydantic_model(
        self,
        model_class: Type[BaseModel],
        data: Dict[str, Any],
    ) -> Tuple[Optional[BaseModel], List[str]]:
        """
        使用Pydantic模型验证数据
        
        Args:
            model_class: Pydantic模型类
            data: 数据字典
            
        Returns:
            (验证后的模型实例, 错误消息列表)
        """
        errors = []
        try:
            model = model_class(**data)
            return model, errors
        except ValidationError as e:
            for error in e.errors():
                field = ".".join(str(loc) for loc in error["loc"])
                errors.append(f"{field}: {error['msg']}")
            return None, errors
    
    def fill_missing_fields(
        self,
        parser,  # ExcelParser实例
        output_path: Optional[str] = None,
        fill_value: str = "待补充",
    ) -> Dict[str, int]:
        """
        填充缺失字段（用默认值）
        
        Args:
            parser: ExcelParser实例
            output_path: 输出路径
            fill_value: 填充值
            
        Returns:
            填充字段统计 {sheet_name: filled_count}
        """
        from openpyxl import load_workbook
        
        wb = load_workbook(parser.file_path)
        filled_stats = {}
        
        for sheet_name, required_fields in self.REQUIRED_FIELDS.items():
            if sheet_name not in wb.sheetnames:
                continue
            
            sheet = wb[sheet_name]
            filled_count = 0
            
            for row in sheet.iter_rows(min_row=2):
                field_name = row[0].value
                if field_name and str(field_name).strip() in required_fields:
                    value_cell = row[1]
                    if value_cell.value is None or str(value_cell.value).strip() == "":
                        # 使用自定义默认值或通用值
                        default = self.DEFAULT_VALUES.get(
                            str(field_name).strip(), 
                            fill_value
                        )
                        value_cell.value = default
                        filled_count += 1
            
            filled_stats[sheet_name] = filled_count
        
        # 保存
        save_path = output_path or parser.file_path
        wb.save(save_path)
        wb.close()
        
        return filled_stats


def validate_excel_file(file_path: str) -> ValidationReport:
    """
    便捷函数：验证Excel文件
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        ValidationReport
    """
    from src.services.excel_parser import ExcelParser
    
    parser = ExcelParser(file_path)
    validator = DataValidator()
    report = validator.validate_all(parser)
    parser.close() if hasattr(parser, 'close') else None
    
    return report


# 测试代码
if __name__ == "__main__":
    import sys
    from pathlib import Path
    
    # 测试文件路径
    project_root = Path(__file__).parent.parent.parent
    test_file = project_root / "templates" / "excel_templates" / "项目数据模板.xlsx"
    
    if len(sys.argv) > 1:
        test_file = Path(sys.argv[1])
    
    if not test_file.exists():
        print(f"测试文件不存在: {test_file}")
        sys.exit(1)
    
    print(f"验证文件: {test_file}")
    print("=" * 60)
    
    # 执行验证
    report = validate_excel_file(str(test_file))
    
    # 输出Markdown报告
    print(report.to_markdown())
    
    # 输出JSON
    print("\n" + "=" * 60)
    print("JSON格式:")
    print(report.to_json())