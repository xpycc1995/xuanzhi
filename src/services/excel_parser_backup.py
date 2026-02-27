"""
Excel数据解析服务

负责读取用户填写的Excel模板，解析为Pydantic数据模型。
支持第1章项目概况和第2章选址分析数据的解析。
"""

import os
from typing import Dict, Any, List, Optional, Tuple
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from src.models.project_overview_data import ProjectOverviewData
from src.models.site_selection_data import (
    SiteSelectionData,
    SiteAlternative,
    SiteNaturalConditions,
    SiteExternalConditions,
    SiteSensitiveConditions,
    ConstructionConditions,
    PlanningImpact,
    ConsultationOpinion,
    SchemeComparison,
)
from src.utils.logger import logger


class ExcelParseError(Exception):
    """Excel解析错误"""
    pass


class ExcelParser:
    """
    Excel数据解析器

    解析项目数据Excel文件，支持以下Sheet:
    - 项目基本信息: 第1章数据
    - 备选方案: 第2章备选方案数据
    - 场地条件: 第2章自然条件和外部条件
    - 敏感条件: 第2章敏感条件
    - 施工运营: 第2章施工运营条件
    - 征求意见: 第2章征求意见情况
    - 方案比选: 第2章方案比选数据
    """

    # Sheet名称常量
    SHEET_PROJECT_INFO = "项目基本信息"
    SHEET_ALTERNATIVES = "备选方案"
    SHEET_SITE_CONDITIONS = "场地条件"
    SHEET_SENSITIVE = "敏感条件"
    SHEET_CONSTRUCTION = "施工运营"
    SHEET_CONSULTATION = "征求意见"
    SHEET_COMPARISON = "方案比选"

    def __init__(self, file_path: str):
        """
        初始化解析器

        Args:
            file_path: Excel文件路径
        """
        self.file_path = file_path
        self.workbook: Optional[Workbook] = None
        self._validate_file()

    def _validate_file(self):
        """验证文件是否存在且为Excel格式"""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel文件不存在: {self.file_path}")

        if not self.file_path.lower().endswith(('.xlsx', '.xlsm')):
            raise ExcelParseError(f"不支持的文件格式: {self.file_path}")

    def _load_workbook(self):
        """加载Excel工作簿"""
        if self.workbook is None:
            logger.info(f"加载Excel文件: {self.file_path}")
            self.workbook = load_workbook(self.file_path, data_only=True)
            logger.info(f"工作簿包含Sheet: {self.workbook.sheetnames}")

    def _get_sheet(self, sheet_name: str) -> Optional[Worksheet]:
        """
        获取指定名称的Sheet

        Args:
            sheet_name: Sheet名称

        Returns:
            Worksheet对象，如果不存在则返回None
        """
        self._load_workbook()
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        return None

    def _read_key_value_sheet(self, sheet: Worksheet) -> Dict[str, str]:
        """
        读取键值对格式的Sheet

        格式: 第一列为字段名，第二列为字段值

        Args:
            sheet: Worksheet对象

        Returns:
            键值对字典
        """
        result = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                key = str(row[0]).strip()
                value = str(row[1]).strip() if row[1] is not None else ""
                result[key] = value
        return result

    def _read_category_sheet(self, sheet: Worksheet) -> Dict[str, Dict[str, Any]]:
        """
        读取分类格式的Sheet

        格式: 第一列类别，第二列项目，第三列内容

        Args:
            sheet: Worksheet对象

        Returns:
            分类字典
        """
        result: Dict[str, Dict[str, Any]] = {}
        current_category = None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                # 新类别
                current_category = str(row[0]).strip()
                if current_category not in result:
                    result[current_category] = {}

            if current_category and row[1] is not None:
                item_name = str(row[1]).strip()
                item_value = row[2] if len(row) > 2 and row[2] is not None else ""
                result[current_category][item_name] = item_value

        return result

    def _read_table_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """
        读取表格格式的Sheet

        格式: 第一行为表头，后续行为数据

        Args:
            sheet: Worksheet对象

        Returns:
            数据行列表
        """
        result = []
        headers = []

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                # 第一行为表头
                headers = [str(cell).strip() if cell else "" for cell in row]
            else:
                # 数据行
                row_data = {}
                for j, cell in enumerate(row):
                    if j < len(headers) and headers[j]:
                        row_data[headers[j]] = cell
                if row_data:
                    result.append(row_data)

        return result

    def parse_project_overview(self) -> ProjectOverviewData:
        """
        解析第1章：项目基本信息

        Returns:
            ProjectOverviewData对象
        """
        logger.info("解析项目基本信息...")

        sheet = self._get_sheet(self.SHEET_PROJECT_INFO)
        if sheet is None:
            raise ExcelParseError(f"缺少Sheet: {self.SHEET_PROJECT_INFO}")

        data = self._read_key_value_sheet(sheet)
        logger.info(f"读取到 {len(data)} 个项目信息字段")

        try:
            project_data = ProjectOverviewData(
                项目名称=data.get("项目名称", ""),
                项目代码=data.get("项目代码"),
                建设单位=data.get("建设单位", ""),
                建设性质=data.get("建设性质", ""),
                项目投资=data.get("项目投资", ""),
                项目选址=data.get("项目选址", data.get("项目位置", "")),
                建设内容=data.get("建设内容", ""),
                建设规模=data.get("建设规模"),
                建设期限=data.get("建设期限"),
            )
            logger.info(f"项目基本信息解析成功: {project_data.项目名称}")
            return project_data

        except ValidationError as e:
            raise ExcelParseError(f"项目基本信息数据验证失败: {str(e)}")

    def parse_site_selection(self) -> SiteSelectionData:
        """
        解析第2章：选址分析数据

        Returns:
            SiteSelectionData对象
        """
        logger.info("解析选址分析数据...")

        # 解析各个部分
        project_info = self._parse_project_info_for_site()
        alternatives = self._parse_alternatives()
        site_conditions = self._parse_site_conditions()
        sensitive_conditions = self._parse_sensitive_conditions()
        construction_conditions = self._parse_construction_conditions()
        planning_impact = self._parse_planning_impact()
        consultation_opinions = self._parse_consultation_opinions()
        comparison = self._parse_scheme_comparison()
        principles = self._parse_site_principles()

        try:
            site_data = SiteSelectionData(
                项目基本信息=project_info,
                选址原则=principles,
                备选方案=alternatives,
                场地自然条件=site_conditions["自然条件"],
                外部配套条件=site_conditions["外部配套"],
                选址敏感条件=sensitive_conditions,
                施工运营条件=construction_conditions,
                规划影响=planning_impact,
                征求意见情况=consultation_opinions,
                方案比选=comparison,
            )
            logger.info("选址分析数据解析成功")
            return site_data

        except ValidationError as e:
            raise ExcelParseError(f"选址分析数据验证失败: {str(e)}")

    def _parse_project_info_for_site(self) -> Dict[str, str]:
        """解析项目基本信息（用于选址数据）"""
        sheet = self._get_sheet(self.SHEET_PROJECT_INFO)
        if sheet:
            return self._read_key_value_sheet(sheet)
        return {}

    def _parse_alternatives(self) -> List[SiteAlternative]:
        """解析备选方案"""
        logger.info("  解析备选方案...")
        sheet = self._get_sheet(self.SHEET_ALTERNATIVES)

        if sheet is None:
            raise ExcelParseError(f"缺少Sheet: {self.SHEET_ALTERNATIVES}")

        rows = self._read_table_sheet(sheet)
        alternatives = []

        for row in rows:
            try:
                # 获取面积值（支持多种列名格式）
                面积值 = row.get("面积") or row.get("面积(平方米)") or row.get("面积（平方米）") or 0
                if 面积值:
                    面积值 = float(面积值) if 面积值 else 0

                # 构建四至范围
                四至范围 = {
                    "东": row.get("东", row.get("东侧", "")),
                    "南": row.get("南", row.get("南侧", "")),
                    "西": row.get("西", row.get("西侧", "")),
                    "北": row.get("北", row.get("北侧", "")),
                }

                # 构建土地利用现状
                土地利用现状 = {}
                for key in ["农村道路", "林地", "园地", "耕地", "建设用地", "交通运输用地", "农村宅基地"]:
                    if key in row and row[key]:
                        土地利用现状[key] = str(row[key])

                alt = SiteAlternative(
                    方案编号=str(row.get("方案编号", "")),
                    方案名称=row.get("方案名称", ""),
                    位置=row.get("位置", ""),
                    面积=面积值,
                    四至范围=四至范围,
                    土地利用现状=土地利用现状 if 土地利用现状 else {"未分类": "0"},
                    是否占用耕地=self._parse_bool(row.get("是否占用耕地", "否")),
                    是否占用永久基本农田=self._parse_bool(row.get("是否占用永久基本农田", "否")),
                    是否涉及未利用地=self._parse_bool(row.get("是否涉及未利用地", "否")),
                    建设内容=row.get("建设内容", ""),
                    工艺流程=row.get("工艺流程"),
                    出水标准=row.get("出水标准"),
                )
                alternatives.append(alt)

            except Exception as e:
                logger.warning(f"解析备选方案行失败: {str(e)}")
                continue

        if len(alternatives) < 2:
            raise ExcelParseError(f"至少需要2个备选方案，当前解析到{len(alternatives)}个")

        logger.info(f"  解析到 {len(alternatives)} 个备选方案")
        return alternatives

    def _parse_site_conditions(self) -> Dict[str, Any]:
        """解析场地条件"""
        logger.info("  解析场地条件...")
        sheet = self._get_sheet(self.SHEET_SITE_CONDITIONS)

        result = {
            "自然条件": None,
            "外部配套": None,
        }

        if sheet is None:
            raise ExcelParseError(f"缺少Sheet: {self.SHEET_SITE_CONDITIONS}")

        categories = self._read_category_sheet(sheet)

        # 构建自然条件
        natural_conditions = SiteNaturalConditions(
            地形地貌=categories.get("地形地貌", {}),
            气候=categories.get("气候", {}),
            区域地质构造=categories.get("区域地质构造", categories.get("地质构造", {})),
            水文地质条件=categories.get("水文地质条件", categories.get("水文地质", {})),
            工程地质=categories.get("工程地质", {}),
            地震=categories.get("地震", {}),
        )
        result["自然条件"] = natural_conditions

        # 构建外部配套条件
        external = categories.get("外部配套条件", categories.get("外部配套", {}))
        result["外部配套"] = SiteExternalConditions(
            周边建筑物=external.get("周边建筑物", ""),
            供水=external.get("供水", ""),
            供电=external.get("供电", ""),
            通讯=external.get("通讯", ""),
            交通=external.get("交通", ""),
            建材来源=external.get("建材来源", ""),
            是否压覆文物=self._parse_bool(external.get("是否压覆文物", "否")),
            是否影响防洪=self._parse_bool(external.get("是否影响防洪", "否")),
        )

        logger.info("  场地条件解析成功")
        return result

    def _parse_sensitive_conditions(self) -> SiteSensitiveConditions:
        """解析敏感条件"""
        logger.info("  解析敏感条件...")
        sheet = self._get_sheet(self.SHEET_SENSITIVE)

        if sheet is None:
            logger.warning(f"缺少Sheet: {self.SHEET_SENSITIVE}，使用默认值")
            return SiteSensitiveConditions(
                历史保护={},
                生态保护={},
                矿产资源={},
                安全防护={},
                重要设施={},
                耕地和基本农田={},
                生态保护红线={},
            )

        categories = self._read_category_sheet(sheet)

        # 处理布尔值
        def to_bool_dict(d: Dict) -> Dict[str, bool]:
            return {k: self._parse_bool(v) for k, v in d.items()}

        def to_str_dict(d: Dict) -> Dict[str, str]:
            return {k: str(v) if v is not None else "" for k, v in d.items()}

        return SiteSensitiveConditions(
            历史保护=to_bool_dict(categories.get("历史保护", {})),
            生态保护=to_str_dict(categories.get("生态保护", {})),
            矿产资源=to_bool_dict(categories.get("矿产资源", {})),
            安全防护=to_bool_dict(categories.get("安全防护", {})),
            重要设施=to_str_dict(categories.get("重要设施", {})),
            耕地和基本农田=to_bool_dict(categories.get("耕地和基本农田", {})),
            生态保护红线=to_bool_dict(categories.get("生态保护红线", {})),
        )

    def _parse_construction_conditions(self) -> ConstructionConditions:
        """解析施工运营条件"""
        logger.info("  解析施工运营条件...")
        sheet = self._get_sheet(self.SHEET_CONSTRUCTION)

        if sheet is None:
            raise ExcelParseError(f"缺少Sheet: {self.SHEET_CONSTRUCTION}")

        data = self._read_key_value_sheet(sheet)

        return ConstructionConditions(
            方案一总投资=data.get("方案一总投资", ""),
            方案二总投资=data.get("方案二总投资", ""),
            政府支持=data.get("政府支持", ""),
            群众支持=data.get("群众支持", ""),
            征地拆迁=data.get("征地拆迁"),
            施工难度=data.get("施工难度", ""),
            材料供应=data.get("材料供应", ""),
        )

    def _parse_planning_impact(self) -> PlanningImpact:
        """解析规划影响"""
        logger.info("  解析规划影响...")
        sheet = self._get_sheet(self.SHEET_CONSTRUCTION)

        if sheet is None:
            # 使用默认值
            return PlanningImpact(
                是否符合国土空间总体规划=True,
                是否列入重点项目库=True,
                重点项目库名称="",
                对区域发展作用="",
            )

        data = self._read_key_value_sheet(sheet)

        # 尝试从施工运营Sheet读取规划影响相关字段
        return PlanningImpact(
            是否符合国土空间总体规划=self._parse_bool(data.get("是否符合国土空间总体规划", "是")),
            是否列入重点项目库=self._parse_bool(data.get("是否列入重点项目库", "是")),
            重点项目库名称=data.get("重点项目库名称"),
            对区域发展作用=data.get("对区域发展作用", ""),
        )

    def _parse_consultation_opinions(self) -> List[ConsultationOpinion]:
        """解析征求意见情况"""
        logger.info("  解析征求意见情况...")
        sheet = self._get_sheet(self.SHEET_CONSULTATION)

        if sheet is None:
            raise ExcelParseError(f"缺少Sheet: {self.SHEET_CONSULTATION}")

        rows = self._read_table_sheet(sheet)
        opinions = []

        for row in rows:
            try:
                opinion = ConsultationOpinion(
                    部门=row.get("部门", ""),
                    日期=str(row.get("日期", "")),
                    复函标题=row.get("复函标题", ""),
                    结论=row.get("结论", ""),
                )
                opinions.append(opinion)
            except Exception as e:
                logger.warning(f"解析征求意见行失败: {str(e)}")
                continue

        if len(opinions) < 3:
            raise ExcelParseError(f"至少需要3个部门意见，当前解析到{len(opinions)}个")

        logger.info(f"  解析到 {len(opinions)} 条征求意见")
        return opinions

    def _parse_scheme_comparison(self) -> SchemeComparison:
        """解析方案比选"""
        logger.info("  解析方案比选...")
        sheet = self._get_sheet(self.SHEET_COMPARISON)

        if sheet is None:
            raise ExcelParseError(f"缺少Sheet: {self.SHEET_COMPARISON}")

        data = self._read_key_value_sheet(sheet)

        # 解析比选因子（可能是逗号分隔的字符串或列表）
        factors_str = data.get("比选因子", "")
        if isinstance(factors_str, str):
            factors = [f.strip() for f in factors_str.split(",") if f.strip()]
        else:
            factors = []

        # 解析推荐理由
        reasons_str = data.get("推荐理由", "")
        if isinstance(reasons_str, str):
            reasons = [r.strip() for r in reasons_str.split(",") if r.strip()]
        else:
            reasons = []

        return SchemeComparison(
            比选因子=factors if factors else ["场地自然条件", "外部配套条件", "选址敏感条件", "施工运营条件", "规划影响条件"],
            推荐方案=data.get("推荐方案", "方案一"),
            推荐理由=reasons if reasons else ["投资较低"],
        )

    def _parse_site_principles(self) -> List[str]:
        """解析选址原则"""
        sheet = self._get_sheet(self.SHEET_PROJECT_INFO)

        if sheet is None:
            return [
                "符合规划要求",
                "不占优质耕地",
                "尽量不迁移民",
                "避免敏感区域",
                "基础设施优先",
                "集约节约利用",
                "方便施工运营",
                "安全可靠",
            ]

        data = self._read_key_value_sheet(sheet)
        principles_str = data.get("选址原则", "")

        if isinstance(principles_str, str) and principles_str:
            return [p.strip() for p in principles_str.split(",") if p.strip()]

        return [
            "符合规划要求",
            "不占优质耕地",
            "尽量不迁移民",
            "避免敏感区域",
            "基础设施优先",
        ]

    def _parse_bool(self, value: Any) -> bool:
        """
        解析布尔值

        Args:
            value: 输入值

        Returns:
            布尔值
        """
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ("是", "true", "yes", "1", "√")
        if isinstance(value, (int, float)):
            return bool(value)
        return False

    def parse_all(self) -> Tuple[ProjectOverviewData, SiteSelectionData]:
        """
        解析所有数据

        Returns:
            (ProjectOverviewData, SiteSelectionData) 元组
        """
        logger.info(f"开始解析Excel文件: {self.file_path}")

        project_overview = self.parse_project_overview()
        site_selection = self.parse_site_selection()

        logger.info("Excel文件解析完成")
        return project_overview, site_selection

    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None


def parse_excel_data(file_path: str) -> Tuple[ProjectOverviewData, SiteSelectionData]:
    """
    便捷函数：解析Excel数据文件

    Args:
        file_path: Excel文件路径

    Returns:
        (ProjectOverviewData, SiteSelectionData) 元组
    """
    parser = ExcelParser(file_path)
    try:
        return parser.parse_all()
    finally:
        parser.close()


# 测试代码
if __name__ == "__main__":
    import sys

    # 查找模板文件
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    template_path = os.path.join(project_root, "templates", "excel_templates", "项目数据模板.xlsx")

    if len(sys.argv) > 1:
        template_path = sys.argv[1]

    print(f"测试Excel解析器...")
    print(f"模板路径: {template_path}")

    if not os.path.exists(template_path):
        print(f"模板文件不存在，请先创建模板文件")
        print(f"预期路径: {template_path}")
        sys.exit(0)

    try:
        parser = ExcelParser(template_path)

        # 测试解析项目基本信息
        print("\n=== 测试解析项目基本信息 ===")
        project_data = parser.parse_project_overview()
        print(f"项目名称: {project_data.项目名称}")
        print(f"建设单位: {project_data.建设单位}")
        print(f"项目投资: {project_data.项目投资}")

        # 测试解析选址数据
        print("\n=== 测试解析选址数据 ===")
        site_data = parser.parse_site_selection()
        print(f"备选方案数: {len(site_data.备选方案)}")
        print(f"征求意见数: {len(site_data.征求意见情况)}")
        print(f"推荐方案: {site_data.方案比选.推荐方案}")

        parser.close()
        print("\n测试通过!")

    except Exception as e:
        print(f"测试失败: {str(e)}")
        import traceback
        traceback.print_exc()