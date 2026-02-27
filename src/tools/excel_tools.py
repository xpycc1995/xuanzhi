"""
Excel智能体工具模块

提供三个核心FunctionTool:
- read_excel: 读取Excel模板字段
- search_knowledge_base: 检索知识库
- write_excel: 写入Excel字段
"""

import json
import logging
from typing import Dict, List, Any, Optional
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# =============================================================================
# Tool 1: read_excel - 读取Excel模板
# =============================================================================

def read_excel(file_path: str, sheet_name: str = "项目基本信息") -> str:
    """
    读取Excel文件指定Sheet的内容
    
    Args:
        file_path: Excel文件路径
        sheet_name: Sheet名称，默认"项目基本信息"
        
    Returns:
        JSON字符串，包含字段列表和空白字段
    """
    try:
        path = Path(file_path)
        if not path.exists():
            return json.dumps({"error": f"文件不存在: {file_path}"}, ensure_ascii=False)
        
        workbook = load_workbook(file_path, data_only=True)
        
        if sheet_name not in workbook.sheetnames:
            return json.dumps({
                "error": f"Sheet不存在: {sheet_name}",
                "available_sheets": workbook.sheetnames
            }, ensure_ascii=False)
        
        sheet = workbook[sheet_name]
        
        # 读取键值对格式
        fields = []
        empty_fields = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                field_name = str(row[0]).strip()
                field_value = str(row[1]).strip() if row[1] is not None else ""
                
                fields.append({
                    "name": field_name,
                    "value": field_value,
                    "is_empty": not field_value
                })
                
                if not field_value:
                    empty_fields.append(field_name)
        
        result = {
            "sheet_name": sheet_name,
            "total_fields": len(fields),
            "empty_fields_count": len(empty_fields),
            "fields": fields,
            "empty_fields": empty_fields
        }
        
        logger.info(f"read_excel完成: {sheet_name}, 共{len(fields)}字段, 空白{len(empty_fields)}个")
        
        return json.dumps(result, ensure_ascii=False, indent=2)
        
    except Exception as e:
        logger.error(f"read_excel失败: {str(e)}")
        return json.dumps({"error": str(e)}, ensure_ascii=False)


def read_excel_all_sheets(file_path: str) -> str:
    """
    读取Excel所有Sheet的内容概要
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        JSON字符串，包含所有Sheet的概要
    """
    try:
        path = Path(file_path)
        if not path.exists():
            return json.dumps({"error": f"文件不存在: {file_path}"}, ensure_ascii=False)
        
        workbook = load_workbook(file_path, data_only=True)
        
        sheets_info = []
        total_empty = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 统计行数和空白字段
            row_count = 0
            empty_count = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    row_count += 1
                    if row[1] is None or str(row[1]).strip() == "":
                        empty_count += 1
            
            sheets_info.append({
                "name": sheet_name,
                "fields_count": row_count,
                "empty_count": empty_count
            })
            total_empty += empty_count
        
        result = {
            "file_name": path.name,
            "sheets_count": len(workbook.sheetnames),
            "total_empty_fields": total_empty,
            "sheets": sheets_info
        }
        
        return json.dumps(result, ensure_ascii=False, indent=2)
        
    except Exception as e:
        return json.dumps({"error": str(e)}, ensure_ascii=False)


# =============================================================================
# Tool 2: search_knowledge_base - 检索知识库
# =============================================================================

def search_knowledge_base(
    query: str,
    n_results: int = 5,
    threshold: float = 0.7
) -> str:
    """
    从RAG知识库检索相关信息
    
    Args:
        query: 查询文本
        n_results: 返回结果数量，默认5
        threshold: 相似度阈值，默认0.7
        
    Returns:
        JSON字符串，包含检索结果
    """
    try:
        from src.rag import get_retriever
        
        retriever = get_retriever()
        
        # 执行检索
        results = retriever.search(
            query=query,
            n_results=n_results,
            threshold=threshold
        )
        
        if not results:
            return json.dumps({
                "query": query,
                "found": False,
                "message": "未找到相关知识",
                "results": []
            }, ensure_ascii=False)
        
        # 格式化结果
        formatted_results = []
        for r in results:
            formatted_results.append({
                "content": r.content,
                "score": round(r.score, 3),
                "source": r.metadata.get("source", "unknown")
            })
        
        result = {
            "query": query,
            "found": True,
            "count": len(results),
            "results": formatted_results
        }
        
        logger.info(f"search_knowledge_base完成: query='{query[:30]}...', found={len(results)}")
        
        return json.dumps(result, ensure_ascii=False, indent=2)
        
    except ImportError:
        return json.dumps({
            "error": "RAG模块未安装或配置错误",
            "query": query
        }, ensure_ascii=False)
    except Exception as e:
        logger.error(f"search_knowledge_base失败: {str(e)}")
        return json.dumps({"error": str(e), "query": query}, ensure_ascii=False)


def search_knowledge_base_for_field(
    field_name: str,
    project_context: str = ""
) -> str:
    """
    为特定字段检索知识库
    
    Args:
        field_name: 字段名称
        project_context: 项目上下文信息
        
    Returns:
        JSON字符串，包含建议值
    """
    # 构建查询
    query = f"{field_name}"
    if project_context:
        query = f"{project_context} {field_name}"
    
    # 检索
    result_json = search_knowledge_base(query, n_results=3, threshold=0.6)
    
    try:
        result = json.loads(result_json)
        
        if result.get("found"):
            # 提取建议值
            suggestions = []
            for r in result.get("results", []):
                content = r.get("content", "")
                if content:
                    suggestions.append(content[:200])  # 截取前200字符
            
            return json.dumps({
                "field_name": field_name,
                "suggestions": suggestions,
                "confidence": max([r.get("score", 0) for r in result.get("results", [])]) if suggestions else 0
            }, ensure_ascii=False)
        else:
            return json.dumps({
                "field_name": field_name,
                "suggestions": [],
                "message": "未找到相关建议"
            }, ensure_ascii=False)
            
    except json.JSONDecodeError:
        return json.dumps({
            "field_name": field_name,
            "error": "解析检索结果失败"
        }, ensure_ascii=False)


# =============================================================================
# Tool 3: write_excel - 写入Excel字段
# =============================================================================

def write_excel(
    file_path: str,
    sheet_name: str,
    field_name: str,
    field_value: str
) -> str:
    """
    写入Excel字段值
    
    Args:
        file_path: Excel文件路径
        sheet_name: Sheet名称
        field_name: 字段名称
        field_value: 字段值
        
    Returns:
        JSON字符串，包含操作结果
    """
    try:
        path = Path(file_path)
        if not path.exists():
            return json.dumps({"error": f"文件不存在: {file_path}"}, ensure_ascii=False)
        
        # 加载工作簿
        workbook = load_workbook(file_path)
        
        if sheet_name not in workbook.sheetnames:
            return json.dumps({
                "error": f"Sheet不存在: {sheet_name}",
                "available_sheets": workbook.sheetnames
            }, ensure_ascii=False)
        
        sheet = workbook[sheet_name]
        
        # 查找字段位置
        field_row = None
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is not None and str(row[0]).strip() == field_name:
                field_row = i
                break
        
        if field_row is None:
            return json.dumps({
                "error": f"字段不存在: {field_name}",
                "sheet": sheet_name
            }, ensure_ascii=False)
        
        # 写入值
        sheet.cell(row=field_row, column=2, value=field_value)
        
        # 保存
        workbook.save(file_path)
        workbook.close()
        
        logger.info(f"write_excel完成: {sheet_name}.{field_name} = {field_value[:50]}")
        
        return json.dumps({
            "success": True,
            "sheet": sheet_name,
            "field": field_name,
            "value": field_value,
            "message": "写入成功"
        }, ensure_ascii=False)
        
    except Exception as e:
        logger.error(f"write_excel失败: {str(e)}")
        return json.dumps({"error": str(e)}, ensure_ascii=False)


def write_excel_batch(
    file_path: str,
    updates: List[Dict[str, Any]]
) -> str:
    """
    批量写入Excel字段
    
    Args:
        file_path: Excel文件路径
        updates: 更新列表，每项包含 {sheet, field, value}
        
    Returns:
        JSON字符串，包含批量操作结果
    """
    try:
        path = Path(file_path)
        if not path.exists():
            return json.dumps({"error": f"文件不存在: {file_path}"}, ensure_ascii=False)
        
        workbook = load_workbook(file_path)
        
        results = []
        success_count = 0
        fail_count = 0
        
        for update in updates:
            sheet_name = update.get("sheet", "项目基本信息")
            field_name = update.get("field")
            field_value = update.get("value", "")
            
            if not field_name:
                results.append({
                    "field": "未知",
                    "success": False,
                    "error": "字段名缺失"
                })
                fail_count += 1
                continue
            
            # 检查Sheet
            if sheet_name not in workbook.sheetnames:
                results.append({
                    "field": field_name,
                    "success": False,
                    "error": f"Sheet不存在: {sheet_name}"
                })
                fail_count += 1
                continue
            
            sheet = workbook[sheet_name]
            
            # 查找字段
            field_row = None
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] is not None and str(row[0]).strip() == field_name:
                    field_row = i
                    break
            
            if field_row:
                sheet.cell(row=field_row, column=2, value=field_value)
                results.append({
                    "field": field_name,
                    "sheet": sheet_name,
                    "success": True
                })
                success_count += 1
            else:
                results.append({
                    "field": field_name,
                    "success": False,
                    "error": "字段不存在"
                })
                fail_count += 1
        
        # 保存
        workbook.save(file_path)
        workbook.close()
        
        logger.info(f"write_excel_batch完成: 成功{success_count}, 失败{fail_count}")
        
        return json.dumps({
            "success": True,
            "total": len(updates),
            "success_count": success_count,
            "fail_count": fail_count,
            "results": results
        }, ensure_ascii=False, indent=2)
        
    except Exception as e:
        return json.dumps({"error": str(e)}, ensure_ascii=False)


# =============================================================================
# 工具注册 (用于AutoGen FunctionTool)
# =============================================================================

def get_excel_tools():
    """
    获取Excel智能体的工具列表
    
    Returns:
        工具函数列表
    """
    return [
        read_excel,
        read_excel_all_sheets,
        search_knowledge_base,
        search_knowledge_base_for_field,
        write_excel,
        write_excel_batch,
    ]


# 工具描述 (用于LLM理解)
TOOL_DESCRIPTIONS = {
    "read_excel": {
        "description": "读取Excel模板文件，获取字段列表和空白字段",
        "parameters": {
            "file_path": "Excel文件路径",
            "sheet_name": "Sheet名称，默认'项目基本信息'"
        }
    },
    "search_knowledge_base": {
        "description": "从RAG知识库检索相关信息",
        "parameters": {
            "query": "查询文本",
            "n_results": "返回结果数量，默认5",
            "threshold": "相似度阈值，默认0.7"
        }
    },
    "write_excel": {
        "description": "写入Excel字段值",
        "parameters": {
            "file_path": "Excel文件路径",
            "sheet_name": "Sheet名称",
            "field_name": "字段名称",
            "field_value": "字段值"
        }
    }
}


if __name__ == "__main__":
    # 测试
    print("Excel工具模块测试")
    print("=" * 50)
    
    # 测试read_excel
    test_result = read_excel("nonexistent.xlsx")
    print(f"\nread_excel测试 (不存在文件):\n{test_result}")
    
    # 测试search_knowledge_base
    test_result = search_knowledge_base("城乡规划", n_results=2)
    print(f"\nsearch_knowledge_base测试:\n{test_result[:200]}...")